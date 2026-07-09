#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CKP Manual Report — No.6 External OS&D Balance by Size 생성기
==============================================================
SCR-005 External OS&D Balance(P_MSPQ38100S_Q) OCI 재현. STATUS=BALANCE(=REQUEST-INCOMING) by size.
원본 3-4 시트 양식: Line | Date | Comp | Model | Style | Color | Type | Supply Plant | STATUS | TOTAL | 사이즈(1~17)
입력 tidy CSV: LINE,OSND_DATE,CMP,MODEL,STYLE_DISP,COLOR_CD,TYPE,SUPPLY_PLANT,STATUS,SIZE_CD,QTY
사용: python osnd_bysize.py OUT.xlsx "시트명" tidy.csv ["타이틀"]
"""
import sys, csv, io, re
from collections import OrderedDict, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SIZES = ["1","1T","2","2T","3","3T","4","4T","5","5T","6","6T","7","7T","8","8T","9","9T",
         "10","10T","11","11T","12","12T","13","13T","14","14T","15","15T","16","16T","17"]  # 원본 3-4 고정 33개
FIXED = ["Line","Date","Comp","Model","Style","Color","Type","Supply Plant","STATUS","TOTAL"]
NAVY="1F3A5F"; ALT="EAF1F8"; GT="D9E1F2"
thin=Side(style="thin",color="BBBBBB"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(sz=9,b=False,col="222222"): return Font(name="Malgun Gothic",size=sz,bold=b,color=col)
HDRF=F(9,True,"FFFFFF")

def _num(v):
    try: return int(round(float(v)))
    except (TypeError,ValueError): return None

def read_tidy(path):
    with open(path,encoding="utf-8") as f: txt=f.read()
    rows=[]
    for d in csv.DictReader(io.StringIO(txt)):
        d={(k or "").strip().upper():v for k,v in d.items()}
        q=_num(d.get("QTY"))
        if not q: continue
        rows.append({"line":(d.get("LINE") or "").strip(),
                     "date":str(d.get("OSND_DATE") or "").strip(),
                     "comp":(d.get("CMP") or "").strip(),
                     "model":(d.get("MODEL") or "").strip(),
                     "style":(d.get("STYLE_DISP") or d.get("STYLE_CD") or "").strip(),
                     "color":(d.get("COLOR_CD") or "").strip(),
                     "type":(d.get("TYPE") or "").strip(),
                     "sup":(d.get("SUPPLY_PLANT") or "").strip(),
                     "status":(d.get("STATUS") or "BALANCE").strip(),
                     "size":(d.get("SIZE_CD") or "").strip(),"qty":q})
    return rows

def _fmt_date(d):
    if re.match(r"^\d{8}$",d):
        m=["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        return f"{int(d[6:8])}-{m[int(d[4:6])]}"
    return d

def build(ws, rows):
    for c,h in enumerate(FIXED,1):
        cell=ws.cell(1,c,h); cell.font=HDRF; cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BORDER
    for i,s in enumerate(SIZES):
        cell=ws.cell(1,11+i,s); cell.font=HDRF; cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="center"); cell.border=BORDER
    sizecol={s:11+i for i,s in enumerate(SIZES)}
    NC=10+len(SIZES)
    key=lambda r:(r["line"],r["date"],r["comp"],r["model"],r["style"],r["color"],r["type"],r["sup"],r["status"])
    grouped=OrderedDict()
    for r in rows: grouped.setdefault(key(r),defaultdict(int))[r["size"]]+=r["qty"]
    ordered=sorted(grouped.items(),key=lambda kv:(kv[0][0],kv[0][1],kv[0][4]))
    # G-Total (row2)
    for c in range(1,NC+1):
        ws.cell(2,c).fill=PatternFill("solid",fgColor=GT); ws.cell(2,c).border=BORDER; ws.cell(2,c).font=F(9,True)
    ws.cell(2,9,"G-Total").font=F(9,True)
    gsz=defaultdict(int); gtot=0; rr=3; alt=0
    for (k,szm) in ordered:
        line,date,comp,model,style,color,typ,sup,status=k
        tot=sum(szm.values()); alt^=1
        base=[line,_fmt_date(date),comp,model,style,color,typ,sup,status,tot]
        for c,v in enumerate(base,1):
            cell=ws.cell(rr,c,v); cell.font=F(9); cell.border=BORDER
            if alt: cell.fill=PatternFill("solid",fgColor=ALT)
        for s,cc in sizecol.items():
            v=szm.get(s,0); cell=ws.cell(rr,cc,v or None); cell.font=F(9); cell.border=BORDER
            if alt: cell.fill=PatternFill("solid",fgColor=ALT)
            gsz[s]+=v
        gtot+=tot; rr+=1
    ws.cell(2,10,gtot).font=F(9,True)
    for s,cc in sizecol.items():
        if gsz.get(s): ws.cell(2,cc,gsz[s]).font=F(9,True)
    ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=8
    ws.column_dimensions["C"].width=6; ws.column_dimensions["D"].width=20
    ws.column_dimensions["E"].width=20; ws.column_dimensions["F"].width=14
    ws.column_dimensions["G"].width=14; ws.column_dimensions["H"].width=11
    ws.column_dimensions["I"].width=9; ws.column_dimensions["J"].width=8
    for c in range(11,NC+1): ws.column_dimensions[get_column_letter(c)].width=5
    ws.freeze_panes="K3"; ws.sheet_view.showGridLines=False
    ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    return gtot

def main():
    out=sys.argv[1]; sheet=sys.argv[2]; tidy=sys.argv[3]
    rows=read_tidy(tidy)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=sheet[:31]
    g=build(ws,rows); wb.save(out)
    print(f"SAVED: {out} ({sheet}) rows={len(rows)} Balance={g:,}")

if __name__=="__main__": main()
