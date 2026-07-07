#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CKP Manual Report — Balance ... by size 생성기 v2 (원본 양식 준수)
===================================================================
원본 「CKP Manual Report (종합)」의 3-2 by-size 시트 양식을 그대로 재현한다.
- 사이즈 컬럼은 원본 고정 세트(값 없어도 항상 표시). G-Total(맨 위) + (Line,Style) Total 소계.
- IP(No.3·4): 단일 사이즈축 1,1T…18. 컬럼 Line|Model|Style|Item Class|MCS Color|FA Date|Div|Shortage|<사이즈>
- PH in Market(No.11): 양식 다름 — Div 없음, ME/WO 이중 사이즈 헤더(같은 물리열에 남/여 라벨).
  성별(Gen)에 따라 ME열 or WO열에 배치. 상단 타이틀행.

입력 tidy CSV 컬럼: LINE,MODEL_NAME,STYLE_CD,ITEM_CLASS,MCS_COLOR,FA_DATE,DIV,GEN,SIZE_CD,QTY
사용:  python bysize_v2.py ip  OUT.xlsx "시트명" tidy.csv
       python bysize_v2.py ph  OUT.xlsx "시트명" tidy.csv  ["타이틀"]
"""
import sys, csv, io
from collections import OrderedDict, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- 원본에서 추출한 고정 사이즈 라벨 ----
SIZE_IP = ["1","1T","2","2T","3","3T","4","4T","5","5T","6","6T","7","7T","8","8T","9","9T",
           "10","10T","11","11T","12","12T","13","13T","14","14T","15","15T","16","16T","17","17T","18"]  # col 9..43
PH_ME = ["1","1T","2","2T","3","3T","4","4T","5","5T","6","6T","7","7T","8","8T","9","9T",
         "10","10T","11","11T","12","12T","13","13T","14","14T","15"]                                     # col 9..37
PH_WO = ["5","5T","6","6T","7","7T","8","8T","9","9T","10","10T","11","11T","12","12T","13","13T","14","14T","15"]  # col 14..34
PH_ME_C0 = 9    # ME 라벨 시작 컬럼
PH_WO_C0 = 14   # WO 라벨 시작 컬럼

NAVY="1F3A5F"; ALT="EAF1F8"; SUB="FFF2CC"; GT="D9E1F2"
thin=Side(style="thin",color="BBBBBB"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(sz=9,b=False,col="222222"): return Font(name="Malgun Gothic",size=sz,bold=b,color=col)
HDRF=F(9,True,"FFFFFF")

def _num(v):
    try: return int(float(v))
    except (TypeError,ValueError): return None

def read_tidy(path):
    with open(path,encoding="utf-8") as f: txt=f.read()
    rows=[]
    for d in csv.DictReader(io.StringIO(txt)):
        d={(k or "").strip().upper():v for k,v in d.items()}
        q=_num(d.get("QTY"))
        if not q: continue
        rows.append({"line":(d.get("LINE") or "").strip(),
                     "model":(d.get("MODEL_NAME") or "").strip(),
                     "style":(d.get("STYLE_CD") or "").strip(),
                     "ic":(d.get("ITEM_CLASS") or "").strip(),
                     "color":(d.get("MCS_COLOR") or "").strip(),
                     "fa":str(d.get("FA_DATE") or "").strip(),
                     "div":(d.get("DIV") or "").strip(),
                     "gen":(d.get("GEN") or "").strip().upper(),
                     "size":(d.get("SIZE_CD") or "").strip(),"qty":q})
    return rows

def _fmt_fa(d):
    import re
    if re.match(r"^\d{8}$",d):
        m=["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        return f"{int(d[6:8])}-{m[int(d[4:6])]}"
    return d

def _hdr_cell(ws,r,c,val,fill=NAVY):
    cell=ws.cell(r,c,val); cell.font=HDRF; cell.fill=PatternFill("solid",fgColor=fill)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BORDER
    return cell

# =================== IP (단일축) ===================
def build_ip(ws, rows):
    fixed=["Line","Model Name","Style Code","Item Class","MCS Color","FA Date","Div","Shortage"]
    for c,h in enumerate(fixed,1): _hdr_cell(ws,1,c,h)
    for i,s in enumerate(SIZE_IP): _hdr_cell(ws,1,9+i,s)
    sizecol={s:9+i for i,s in enumerate(SIZE_IP)}
    NC=8+len(SIZE_IP)
    key=lambda r:(r["line"],r["model"],r["style"],r["ic"],r["color"],r["fa"],r["div"])
    grouped=OrderedDict()
    for r in rows: grouped.setdefault(key(r),defaultdict(int))[r["size"]]+=r["qty"]
    ordered=sorted(grouped.items(),key=lambda kv:(kv[0][0],kv[0][2],kv[0][5]))
    # G-Total (row2)
    for c in range(1,NC+1):
        ws.cell(2,c).fill=PatternFill("solid",fgColor=GT); ws.cell(2,c).border=BORDER; ws.cell(2,c).font=F(9,True)
    ws.cell(2,7,"G-Total").font=F(9,True)
    gsz=defaultdict(int); gtot=0
    rr=3; cur=None; ls_sz=defaultdict(int); ls_tot=0; alt=0
    def flush():
        nonlocal rr
        if cur is None: return
        ws.cell(rr,7,"Total").font=F(9,True); ws.cell(rr,8,ls_tot).font=F(9,True)
        for s,cc in sizecol.items():
            if ls_sz.get(s): ws.cell(rr,cc,ls_sz[s]).font=F(9,True)
        for c in range(1,NC+1):
            ws.cell(rr,c).fill=PatternFill("solid",fgColor=SUB); ws.cell(rr,c).border=BORDER
    for (k,szm) in ordered:
        line,model,style,ic,color,fa,div=k
        if cur is not None and (line,style)!=cur:
            flush(); rr+=1; cur=None; ls_sz=defaultdict(int); ls_tot=0
        if cur is None: cur=(line,style); ls_sz=defaultdict(int); ls_tot=0; alt^=1
        short=sum(szm.values())
        base=[line,model,style,ic,color,_fmt_fa(fa),div,short]
        for c,v in enumerate(base,1):
            cell=ws.cell(rr,c,v); cell.font=F(9); cell.border=BORDER
            if alt: cell.fill=PatternFill("solid",fgColor=ALT)
        for s,cc in sizecol.items():
            v=szm.get(s,0)
            cell=ws.cell(rr,cc,v or None); cell.font=F(9); cell.border=BORDER
            if alt: cell.fill=PatternFill("solid",fgColor=ALT)
            ls_sz[s]+=v; gsz[s]+=v
        ls_tot+=short; gtot+=short; rr+=1
    flush(); rr+=1
    ws.cell(2,8,gtot).font=F(9,True)
    for s,cc in sizecol.items():
        if gsz.get(s): ws.cell(2,cc,gsz[s]).font=F(9,True)
    _finalize(ws,NC,froze="I3")
    return gtot

# =================== PH in Market (ME/WO 이중) ===================
def build_ph(ws, rows, title="BALANCE IN MARKET"):
    ws.cell(1,1,title).font=F(11,True,NAVY)
    fixed=["Line","Model Name","Style Code","Item Class","MCS Color","FA Date","Shortage"]
    for c,h in enumerate(fixed,1):
        _hdr_cell(ws,2,c,h); ws.merge_cells(start_row=2,start_column=c,end_row=3,end_column=c)
    _hdr_cell(ws,2,8,"ME"); _hdr_cell(ws,3,8,"WO")
    for i,s in enumerate(PH_ME): _hdr_cell(ws,2,PH_ME_C0+i,s)
    for i,s in enumerate(PH_WO): _hdr_cell(ws,3,PH_WO_C0+i,s)
    me_col={s:PH_ME_C0+i for i,s in enumerate(PH_ME)}
    wo_col={s:PH_WO_C0+i for i,s in enumerate(PH_WO)}
    NC=8+len(PH_ME)   # col 9..37
    def gender_col(gen,model,size):
        w = (gen=="WO") or ("(W)" in (model or ""))
        return (wo_col if w else me_col).get(size)
    key=lambda r:(r["line"],r["model"],r["style"],r["ic"],r["color"],r["fa"])
    grouped=OrderedDict()
    for r in rows: grouped.setdefault(key(r),{"g":r["gen"],"sz":defaultdict(int)})["sz"][r["size"]]+=r["qty"]
    ordered=sorted(grouped.items(),key=lambda kv:(kv[0][0],kv[0][2],kv[0][5]))
    # G-Total (row4)
    for c in range(1,NC+1):
        ws.cell(4,c).fill=PatternFill("solid",fgColor=GT); ws.cell(4,c).border=BORDER; ws.cell(4,c).font=F(9,True)
    ws.cell(4,6,"G-Total").font=F(9,True)
    gcolv=defaultdict(int); gtot=0
    rr=5; cur=None; ls_col=defaultdict(int); ls_tot=0; alt=0
    def flush():
        nonlocal rr
        if cur is None: return
        ws.cell(rr,6,"Total").font=F(9,True); ws.cell(rr,7,ls_tot).font=F(9,True)
        for cc,v in ls_col.items():
            if v: ws.cell(rr,cc,v).font=F(9,True)
        for c in range(1,NC+1):
            ws.cell(rr,c).fill=PatternFill("solid",fgColor=SUB); ws.cell(rr,c).border=BORDER
    for (k,info) in ordered:
        line,model,style,ic,color,fa=k; gen=info["g"]; szm=info["sz"]
        if cur is not None and (line,style)!=cur:
            flush(); rr+=1; cur=None; ls_col=defaultdict(int); ls_tot=0
        if cur is None: cur=(line,style); ls_col=defaultdict(int); ls_tot=0; alt^=1
        short=sum(szm.values())
        base=[line,model,style,ic,color,_fmt_fa(fa),short,""]
        for c,v in enumerate(base,1):
            cell=ws.cell(rr,c,v if v!="" else None); cell.font=F(9); cell.border=BORDER
            if alt: cell.fill=PatternFill("solid",fgColor=ALT)
        for c in range(9,NC+1):
            cell=ws.cell(rr,c); cell.border=BORDER
            if alt: cell.fill=PatternFill("solid",fgColor=ALT)
        for s,v in szm.items():
            cc=gender_col(gen,model,s)
            if not cc: continue
            cell=ws.cell(rr,cc,(ws.cell(rr,cc).value or 0)+v); cell.font=F(9); cell.border=BORDER
            if alt: cell.fill=PatternFill("solid",fgColor=ALT)
            ls_col[cc]+=v; gcolv[cc]+=v
        ls_tot+=short; gtot+=short; rr+=1
    flush(); rr+=1
    ws.cell(4,7,gtot).font=F(9,True)
    for cc,v in gcolv.items():
        if v: ws.cell(4,cc,v).font=F(9,True)
    _finalize(ws,NC,froze="I5")
    return gtot

def _finalize(ws,NC,froze):
    ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=22
    ws.column_dimensions["C"].width=13; ws.column_dimensions["D"].width=9
    ws.column_dimensions["E"].width=16; ws.column_dimensions["F"].width=9
    for c in range(7,NC+1): ws.column_dimensions[get_column_letter(c)].width=5.5
    ws.freeze_panes=froze; ws.sheet_view.showGridLines=False
    ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

def main():
    mode=sys.argv[1]; out=sys.argv[2]; sheet=sys.argv[3]; tidy=sys.argv[4]
    title=sys.argv[5] if len(sys.argv)>5 else "BALANCE IN MARKET"
    rows=read_tidy(tidy)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=sheet[:31]
    g = build_ph(ws,rows,title) if mode=="ph" else build_ip(ws,rows)
    wb.save(out); print(f"SAVED: {out}  ({sheet}) rows={len(rows)} Shortage={g:,}")

if __name__=="__main__": main()
