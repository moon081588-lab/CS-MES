# -*- coding: utf-8 -*-
import sys, csv, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
SIZES=["1","1T","2","2T","3","3T","4","4T","5","5T","6","6T","7","7T","8","8T","9","9T","10","10T","11","11T","12","12T","13","13T","14","14T","15","15T","16","16T","17"]
FIXED=["Line","Date","Comp","Model","Style","Color","Type","Supply Plant","STATUS","TOTAL"]
NAVY="1F3A5F";ALT="EAF1F8";GT="D9E1F2";thin=Side(style="thin",color="BBBBBB");BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def Fn(s=9,b=False,c="222222"):return Font(name="Malgun Gothic",size=s,bold=b,color=c)
import re
def fmtd(d):
    d=str(d).strip()
    if re.match(r"^\d{8}$",d):
        m=["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];return f"{int(d[6:8])}-{m[int(d[4:6])]}"
    return d
def num(v):
    try:return int(round(float(v)))
    except:return 0
rows=list(csv.DictReader(io.StringIO(open(sys.argv[3],encoding="utf-8").read())))
rows=[{k.strip().upper():v for k,v in r.items()} for r in rows if r.get("LINE") and "rows selected" not in str(r.get("LINE"))]
wb=openpyxl.Workbook();ws=wb.active;ws.title=sys.argv[2][:31]
NC=10+len(SIZES)
for c,h in enumerate(FIXED,1):
    x=ws.cell(1,c,h);x.font=Fn(9,True,"FFFFFF");x.fill=PatternFill("solid",fgColor=NAVY);x.border=BORDER;x.alignment=Alignment(horizontal="center",wrap_text=True)
for i,s in enumerate(SIZES):
    x=ws.cell(1,11+i,s);x.font=Fn(9,True,"FFFFFF");x.fill=PatternFill("solid",fgColor=NAVY);x.border=BORDER;x.alignment=Alignment(horizontal="center")
for c in range(1,NC+1):
    ws.cell(2,c).fill=PatternFill("solid",fgColor=GT);ws.cell(2,c).border=BORDER;ws.cell(2,c).font=Fn(9,True)
ws.cell(2,9,"G-Total").font=Fn(9,True)
gtot=0;gsz=[0]*len(SIZES);rr=3;alt=0
for r in rows:
    alt^=1
    base=[r.get("LINE"),fmtd(r.get("DT")),r.get("CMP"),r.get("MODEL"),r.get("STYLE"),r.get("COLORV"),r.get("TYPE"),r.get("SUP"),r.get("STATUSV") or "BALANCE",num(r.get("TOT"))]
    for c,v in enumerate(base,1):
        x=ws.cell(rr,c,v);x.font=Fn(9);x.border=BORDER
        if alt:x.fill=PatternFill("solid",fgColor=ALT)
    for i in range(len(SIZES)):
        v=num(r.get(f"Z{i+1}"));x=ws.cell(rr,11+i,v or None);x.font=Fn(9);x.border=BORDER
        if alt:x.fill=PatternFill("solid",fgColor=ALT)
        gsz[i]+=v
    gtot+=num(r.get("TOT"));rr+=1
ws.cell(2,10,gtot).font=Fn(9,True)
for i in range(len(SIZES)):
    if gsz[i]:ws.cell(2,11+i,gsz[i]).font=Fn(9,True)
for c,w in zip("ABCDEFGHIJ",[8,8,6,20,20,14,14,11,9,8]):ws.column_dimensions[c].width=w
for c in range(11,NC+1):ws.column_dimensions[get_column_letter(c)].width=5
ws.freeze_panes="K3";ws.sheet_view.showGridLines=False
ws.page_setup.orientation="landscape";ws.page_setup.paperSize=9;ws.page_setup.fitToWidth=1;ws.page_setup.fitToHeight=0
ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
wb.save(sys.argv[1]);print(f"SAVED {sys.argv[1]} rows={len(rows)} Balance={gtot:,}")
