#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CKP Manual Report — No.1 DAILY REPORT SCAN (PHH SCAN CTM PHYLON) 생성기
=======================================================================
원본 양식: 고정 6컬럼(LINE·MODEL·COLOR·STYLE CD·MCS·YIELD) + 날짜별 5소컬럼(TARGET·NORMAL·OS&D·TOTAL Prs·TOTAL Kg) + WEEK 소계.
소스: NORMAL = OCI.POP_PCARD_SCAN op_cd='PHH'(파일론 스캔). MCS/COLOR = MSPD_BATCH_PLAN, MODEL = MSBS_ITEM_STYLE.
[미확정 소스라 빈칸] YIELD, TARGET(계획), OS&D(불량), TOTAL Kg(=Prs×YIELD). 확정되면 채운다.
입력 tidy CSV: LINE,MODELV,COLORV,STYLE,MCS,N1..N6,WTOT (N1..N6 = 날짜별 NORMAL)
사용: python daily_scan.py OUT.xlsx tidy.csv "타이틀" "MM/DD,MM/DD,..(6개)"
"""
import sys, csv, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUB=["TARGET","NORMAL","OS&D","TOTAL Prs","TOTAL Kg"]
FIXED=["LINE","MODEL","COLOR","STYLE CD","MCS","YIELD"]
NAVY="1F3A5F";ALT="EAF1F8";thin=Side(style="thin",color="BBBBBB");BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def Fn(s=9,b=False,c="222222"):return Font(name="Malgun Gothic",size=s,bold=b,color=c)
def num(v):
    try:return int(round(float(v)))
    except:return 0

def main():
    out=sys.argv[1]; tidy=sys.argv[2]
    title=sys.argv[3] if len(sys.argv)>3 else "DAILY REPORT SCAN AUTO PHYLON"
    dates=(sys.argv[4].split(",") if len(sys.argv)>4 else ["06/30","07/01","07/02","07/03","07/04","07/06"])
    nd=len(dates)
    rows=[{k.strip().upper():v for k,v in r.items()} for r in csv.DictReader(io.StringIO(open(tidy,encoding="utf-8").read())) if r.get("LINE") and "rows selected" not in str(r.get("LINE"))]
    wb=openpyxl.Workbook();ws=wb.active;ws.title="DAILY REPORT SCAN"
    NC=6+nd*5+2
    ws.cell(1,1,title).font=Fn(12,True,NAVY)
    ws.cell(3,1,"PHH (SCAN CTM PHYLON)").font=Fn(10,True,NAVY)
    # r4/r5 헤더
    for i,h in enumerate(FIXED,1):
        c=ws.cell(4,i,h);c.font=Fn(9,True,"FFFFFF");c.fill=PatternFill("solid",fgColor=NAVY);c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);c.border=BORDER
        ws.merge_cells(start_row=4,start_column=i,end_row=5,end_column=i)
        ws.cell(5,i).fill=PatternFill("solid",fgColor=NAVY);ws.cell(5,i).border=BORDER
    for d in range(nd):
        c0=7+d*5
        hc=ws.cell(4,c0,f"OUT {dates[d]}");hc.font=Fn(9,True,"FFFFFF");hc.fill=PatternFill("solid",fgColor=NAVY);hc.alignment=Alignment(horizontal="center");hc.border=BORDER
        ws.merge_cells(start_row=4,start_column=c0,end_row=4,end_column=c0+4)
        for j,s in enumerate(SUB):
            cc=ws.cell(5,c0+j,s);cc.font=Fn(8,True,"FFFFFF");cc.fill=PatternFill("solid",fgColor=NAVY);cc.alignment=Alignment(horizontal="center",wrap_text=True);cc.border=BORDER
    wk=7+nd*5
    hc=ws.cell(4,wk,"WEEK 1");hc.font=Fn(9,True,"FFFFFF");hc.fill=PatternFill("solid",fgColor=NAVY);hc.alignment=Alignment(horizontal="center");hc.border=BORDER
    ws.merge_cells(start_row=4,start_column=wk,end_row=4,end_column=wk+1)
    for j,s in enumerate(["TOTAL Prs","TOTAL Kg"]):
        cc=ws.cell(5,wk+j,s);cc.font=Fn(8,True,"FFFFFF");cc.fill=PatternFill("solid",fgColor=NAVY);cc.alignment=Alignment(horizontal="center",wrap_text=True);cc.border=BORDER
    # 데이터
    rr=6;alt=0;prev_line=None
    for r in rows:
        alt^=1
        line=r.get("LINE") or ""
        showline = "" if line==prev_line else line
        prev_line=line
        base=[showline, r.get("MODELV") or "", r.get("COLORV") or "", r.get("STYLE") or "", r.get("MCS") or "", ""]  # YIELD 빈칸
        for c,v in enumerate(base,1):
            x=ws.cell(rr,c,v);x.font=Fn(9);x.border=BORDER
            if alt:x.fill=PatternFill("solid",fgColor=ALT)
        for d in range(nd):
            n=num(r.get(f"N{d+1}"));c0=7+d*5
            vals=[None, (n or None), None, (n or None), None]  # TARGET/OS&D/TOTAL Kg 빈칸, NORMAL=Prs=n
            for j,v in enumerate(vals):
                x=ws.cell(rr,c0+j,v);x.font=Fn(9);x.border=BORDER
                if alt:x.fill=PatternFill("solid",fgColor=ALT)
        wtot=num(r.get("WTOT"))
        for j,v in enumerate([wtot or None,None]):
            x=ws.cell(rr,wk+j,v);x.font=Fn(9,True);x.border=BORDER
            if alt:x.fill=PatternFill("solid",fgColor=ALT)
        rr+=1
    ws.column_dimensions["A"].width=7;ws.column_dimensions["B"].width=22;ws.column_dimensions["C"].width=16
    ws.column_dimensions["D"].width=12;ws.column_dimensions["E"].width=11;ws.column_dimensions["F"].width=7
    for c in range(7,NC+1):ws.column_dimensions[get_column_letter(c)].width=7
    ws.freeze_panes="G6";ws.sheet_view.showGridLines=False
    ws.page_setup.orientation="landscape";ws.page_setup.paperSize=9;ws.page_setup.fitToWidth=1;ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    wb.save(out);print(f"SAVED {out} rows={len(rows)}")

if __name__=="__main__": main()
