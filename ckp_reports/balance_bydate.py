#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CKP Manual Report — Balance ... (by date) 생성기  (3-1 계열)
============================================================
검증된 shortage 엔진의 tidy 결과(ITEM_CLASS,FA_WC,STYLE_CD,STYLE_NAME,MCS_COLOR,FA_DATE,QTY)를 받아
원본 "3-1. Balance ..." 시트 레이아웃(FA_DATE 컬럼 PIVOT + PROD Total)으로 xlsx 생성.
헤더: Item Class | FA W/C | Style CD | Style Name | MCS Color | SPRAY | PAD | Total | <날짜들>
(SPRAY/PAD 는 소스 미확인으로 공란)

사용:  python balance_bydate.py OUT.xlsx "SHEET=csv" ["SHEET=csv" ...]
"""
import sys, csv, io, re
from collections import OrderedDict, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3A5F"; ALT="EAF1F8"; SUB="FFF2CC"; GT="D9E1F2"
thin=Side(style="thin", color="BBBBBB"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
HDR=Font(bold=True,color="FFFFFF",size=9,name="Malgun Gothic")
CELL=Font(size=9,name="Malgun Gothic"); BOLD=Font(bold=True,size=9,name="Malgun Gothic")
FIXED=["Item Class","FA W/C","Style CD","Style Name","MCS Color","SPRAY","PAD","Total"]

def _num(v):
    try: return int(float(v))
    except (TypeError,ValueError): return None

def read_tidy(path):
    with open(path,encoding="utf-8") as f: txt=f.read()
    rows=[]
    for d in csv.DictReader(io.StringIO(txt)):
        d={(k or "").strip().upper():v for k,v in d.items()}
        q=_num(d.get("QTY"))
        if not q: continue
        rows.append({"ic":(d.get("ITEM_CLASS") or "").strip(),
                     "wc":(d.get("FA_WC") or d.get("FA_WC_CD") or d.get("LINE") or "").strip(),
                     "style":(d.get("STYLE_CD") or "").strip(),
                     "name":(d.get("STYLE_NAME") or d.get("MODEL_NAME") or "").strip(),
                     "color":(d.get("MCS_COLOR") or "").strip(),
                     "fa":str(d.get("FA_DATE") or "").strip(),"qty":q})
    return rows

def _fmt(d):  # 20260703 -> 07-03
    return d[4:6]+"-"+d[6:8] if re.match(r"^\d{8}$",d) else d

def build_sheet(ws, rows):
    dates=sorted({r["fa"] for r in rows if r["fa"]})
    headers=FIXED+[_fmt(d) for d in dates]
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h); cell.font=HDR; cell.fill=PatternFill("solid",fgColor=NAVY)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BORDER
    key=lambda r:(r["ic"],r["wc"],r["style"],r["name"],r["color"])
    grouped=OrderedDict()
    for r in rows:
        grouped.setdefault(key(r),defaultdict(int))[r["fa"]]+=r["qty"]
    ordered=sorted(grouped.items(),key=lambda kv:(kv[0][1],kv[0][0],kv[0][2]))
    # G-Total placeholder (row 2)
    gt=defaultdict(int); gts=0
    ws.cell(2,7,"G-Total").font=BOLD
    for c in range(1,len(headers)+1):
        ws.cell(2,c).fill=PatternFill("solid",fgColor=GT); ws.cell(2,c).border=BORDER; ws.cell(2,c).font=BOLD
    rr=3; alt=0
    for (k,dm) in ordered:
        ic,wc,style,name,color=k; tot=sum(dm.values())
        vals=[ic,wc,style,name,color,"","",tot]+[dm.get(d,0) or None for d in dates]
        alt^=1
        for c,v in enumerate(vals,1):
            cell=ws.cell(rr,c,v); cell.font=CELL; cell.border=BORDER
            if alt: cell.fill=PatternFill("solid",fgColor=ALT)
        for d in dates: gt[d]+=dm.get(d,0)
        gts+=tot; rr+=1
    ws.cell(2,8,gts).font=BOLD
    for i,d in enumerate(dates): ws.cell(2,9+i,gt.get(d,0) or None).font=BOLD
    widths=[10,8,14,24,18,7,6,10]+[7]*len(dates)
    for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="A3"; ws.sheet_view.showGridLines=False
    ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    return gts

def main():
    out=sys.argv[1]; wb=openpyxl.Workbook(); first=True; summ=[]
    for spec in sys.argv[2:]:
        name,path=spec.split("=",1); rows=read_tidy(path)
        ws=wb.active if first else wb.create_sheet(); first=False; ws.title=name[:31]
        g=build_sheet(ws,rows); summ.append(f"· {name}: {len(rows)}행 / Total {g:,}")
    wb.save(out); print("SAVED:",out); print("\n".join(summ))

if __name__=="__main__": main()
