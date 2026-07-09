#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CKP Manual Report — No.2 "3-1. Balance IP Production (AFTER SCAN UV)" 존(zone) 양식 생성기
============================================================================================
원본 시트가 RED/YELLOW/GREEN 존 범례 + JJ/RJ 이중 D-offset 등 복잡한 커스텀 양식이라,
원본 시트를 그대로 템플릿으로 복사해 데이터행만 비운다(현재 II 생산 부족분 스냅샷).
STOCK/Set Bal/YIELD/Hasil 등 미확정 소스 컬럼은 원본 헤더만 유지.
사용: python no2_zone.py OUT.xlsx SOURCE_WORKBOOK.xlsx
"""
import sys
import openpyxl

SHEET = "3-1. Balance IP Production"
FIRST_DATA = 10   # 헤더 1~9행 유지, 10행부터 데이터

def build(out, src_wb, sheet=SHEET, first_data=FIRST_DATA):
    wb = openpyxl.load_workbook(src_wb)
    for s in list(wb.sheetnames):
        if s != sheet:
            wb.remove(wb[s])
    ws = wb[sheet]
    for mr in [str(m) for m in list(ws.merged_cells.ranges) if m.min_row >= first_data]:
        ws.unmerge_cells(mr)
    maxr, maxc = ws.max_row, ws.max_column
    for r in range(first_data, maxr + 1):
        for c in range(1, maxc + 1):
            ws.cell(r, c).value = None
    if maxr >= first_data:
        ws.delete_rows(first_data, maxr - first_data + 1)
    ws.print_area = None
    wb.save(out)
    print(f"SAVED {out} (원본 존 양식 복사, 데이터 비움)")

if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2])
