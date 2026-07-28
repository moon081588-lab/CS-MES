#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Changshin GMES — No.5 (3-3. Balance IP Outgoing Market) 엔진
============================================================
'출고 부족분(BALANCE OUTGOING MARKET)' 을 원본 엑셀 양식 그대로 만드는 인프로세스 엔진.
컬럼 A~I / 일자열 J~ 작업일 / X=TOTAL · Y=ISSUE · Z=SCAN BEM/BEP · AA=SCAN DI CKP, 동(棟) 세로병합, GRAND TOTAL.

리포트 생성 흐름에서 실제 사용하는 것:
  · load_config    — config.ini 로드(인라인 주석 제거)
  · SHEETS         — 시트 구성(IP/PH/OS)
  · build_buckets  — 날짜 버킷
  · pivot          — SQL 결과(CSV) → 표 데이터 (COLOR/IP SPRAY(BEM)/SCAN DI CKP 채움)
  · build_workbook — report_template.xlsx 를 스타일 도너로 원본 양식 100% 복제
  · color_specs / scan_di_ckp — 선택 후크

[채움 상태] COLOR = MSPD_BATCH_PLAN.MCS_COLOR_CD, SCAN BEM/BEP = POP_PCARD_SCAN(op BEM/BEP).
  IP SPRAY(BEM) = MCS_COLOR_CD (현업 확인: BEM 은 다 색상, COLOR 와 동일) → pivot 에서 자동.
  SCAN DI CKP   = CKP 출고 스캔(MSPD_PCARD_RESULT: PROD·OUT_DATE + MOVE·IN_DATE) → sql.outgoing_market_dickp_sql.
  MOVE 게이트   = sql.outgoing_market_sheet_sql(strict=True) — config [report] strict_outgoing=true 로 켬.
  아직 빈칸(SAP/수기): PAD PRINTING(BEP)=PFC, PH SPRAY(소스 미확인), ISSUE=JIT/plan.
"""
import os, sys, re, configparser, logging, datetime
from core import sql as _BS   # 계열 정의 단일 출처(OM_FAMILIES) 참조

HERE = os.path.dirname(os.path.abspath(__file__))

def load_config(path):
    cp = configparser.ConfigParser(interpolation=None)   # 비밀번호의 % 등 특수문자 안전
    if not cp.read(path, encoding="utf-8"): sys.exit(f"[설정오류] config 없음: {path}")
    # 값 뒤에 붙은 인라인 주석(;) 자동 제거 → 'window_before = 3   ; 설명' 같은 입력도 OK.
    KEEP = {"password", "wallet_password"}
    for sec in cp.sections():
        for k, v in cp[sec].items():
            if k not in KEEP and v and ";" in v:
                cp[sec][k] = v.split(";", 1)[0].strip()
    return cp

def setup_log():
    log=logging.getLogger("bo"); log.setLevel(logging.INFO)
    fmt=logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    sh=logging.StreamHandler(); sh.setFormatter(fmt); log.addHandler(sh)
    try:
        fh=logging.FileHandler(os.path.join(HERE,"balance_outgoing.log"),encoding="utf-8"); fh.setFormatter(fmt); log.addHandler(fh)
    except Exception: pass
    return log
LOG=setup_log()

# [단일 출처] 계열 정의는 sql.OM_FAMILIES 한 곳에서만 관리 → 여기서 파생.
#   (예전엔 여기 SHEETS 와 sql.OM_FAMILIES 에 CP 계열이 이중 정의돼, 한쪽만 고쳐 No.5 CP 제외가 누락됐음.)
SHEETS=[(k, list(v)) for k, v in _BS.OM_FAMILIES.items()]

# ============================================ 선택 후크 (표준 경로는 pivot 이 자동 처리)
def color_specs(plant_cd, style_cd, item_class):
    """COLOR / IP SPRAY(BEM) / PAD PRINTING(BEP) 후크(선택).
    IP SPRAY(BEM) 은 pivot() 에서 MCS_COLOR_CD 로 자동 채우므로 여기서 비워 두면 그 값이 쓰인다.
    PAD PRINTING(BEP) 는 PFC(MUL detail)=SAP 소스 확보 시 세 번째 반환값에 채우면 자동 반영."""
    return ("", "", "")
def scan_di_ckp(plant_cd, fa_wc_cd, style_cd):
    """SCAN DI CKP 행단위 후크(선택). 표준 경로는 sql.outgoing_market_dickp_sql 로 di_map 을 만들어
    pivot(...,di_map) 에 넘긴다. di_map 이 있으면 이 후크는 호출되지 않는다."""
    return ""
# ============================================

def build_buckets(today, before, after):
    """원본과 동일: 작업일(일요일 제외) 기준 D+before … DD … D-after.  반환 (label, yyyymmdd, 일, k)."""
    def walk(n, direction):
        days=[]; d=today
        while len(days)<n:
            d=d+datetime.timedelta(days=direction)
            if d.weekday()!=6: days.append(d)        # 일요일 제외(토요일은 작업일)
        return days
    seq=list(reversed(walk(before,-1)))+[today]+walk(after,+1)
    labels=[f"D+{k}" for k in range(before,0,-1)]+["DD"]+[f"D-{k}" for k in range(1,after+1)]
    ks=list(range(before,0,-1))+[0]+[-k for k in range(1,after+1)]
    return [(labels[i],seq[i].strftime("%Y%m%d"),seq[i].day,ks[i]) for i in range(len(seq))]

def dong(wcg):
    m=re.match(r"[A-Za-z]+",(wcg or "").strip())
    return m.group(0) if m else (wcg or "").strip()

def _clean_color(raw):
    """'OBSIDIAN(45B)' → 'OBSIDIAN', 'NONE'·공란 → ''."""
    c=(raw or "").split("(")[0].strip()
    return "" if c.upper()=="NONE" else c

def pivot(rows, buckets, scan_map, di_map=None):
    """di_map: (plant,line,style)->SCAN DI CKP 수량. 없으면 scan_di_ckp() 후크 사용."""
    bdates=[b[1] for b in buckets]; table={}; colors={}
    for row in rows:
        wcg,plant,ic,line,model,gen,style,fa,qty = row[:9]
        color = row[9] if len(row)>9 else None      # 10번째 컬럼=대표색(있으면)
        dng=dong(wcg) or plant
        key=(dng,plant,ic,line,(model or " ").strip(),(gen or " ").strip(),style)
        table.setdefault(key,{})[fa]=table.setdefault(key,{}).get(fa,0)+(qty or 0)
        if color and style not in colors: colors[style]=color
    out=[]
    for key,dd in table.items():
        cells=[dd.get(bd,0) for bd in bdates]; total=sum(cells)
        if total<=0: continue
        dng,plant,ic,line,model,gen,style=key
        _c,spray,pad=color_specs(plant,style,ic)     # spray/pad 후크(있으면 우선)
        color_full=colors.get(style)
        color=_clean_color(color_full)
        if not spray:                                # IP SPRAY(BEM) = MCS 색상 (현업 확인: 'BEM 은 다 색상, 동일')
            spray=(color_full or "").strip()
        scan_di = di_map.get((plant,line,style),0) if di_map is not None else scan_di_ckp(plant,line,style)
        out.append({"dong":dng,"ic":ic,"line":line,"model":model,"style":style,"color":color,
                    "spray":spray,"pad":pad,"gen":gen,"cells":cells,"total":total,"issue":"",
                    "scan_bem":scan_map.get((plant,line,style),0),"scan_di":scan_di})
    out.sort(key=lambda r:(r["dong"],r["line"],r["style"]))
    return out

# ---------- Excel: 원본을 '스타일 템플릿'으로 사용해 100% 동일 양식 복제 ----------
# 원본(2.3 BALANCE OUTGOING)의 IP0306 시트를 report_template.xlsx(스타일 도너)로 보관하고,
# 각 셀의 폰트·채움·테두리·정렬·표시형식을 그대로 복사한다. 날짜의존 값(날짜/월/일/D오프셋)과
# 본문 데이터만 새로 써넣는다. 색상범례는 원본 이미지를 그대로 잘라낸 legend.png 를 삽입한다.
import os
from copy import copy as _copy
_HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_XLSX = os.path.join(_HERE, "report_template.xlsx")
LEGEND_PNG    = os.path.join(_HERE, "legend.png")

# 템플릿 기준 행: 헤더 1~5, 대표 데이터행 6, 마지막 데이터행 70, GRAND TOTAL 71
T_HDR=range(1,6); T_DATA=6; T_LASTDATA=70; T_GTOT=71; T_NCOL=27

def _copy_style(src, dst):
    if src.has_style:
        dst.font=_copy(src.font); dst.fill=_copy(src.fill)
        dst.border=_copy(src.border); dst.alignment=_copy(src.alignment)
        dst.number_format=src.number_format; dst.protection=_copy(src.protection)

def build_workbook(data_by_sheet, buckets, today, today_str):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    NOFILL=PatternFill(fill_type=None)   # 데이터행은 흰색(원본의 수작업 강조색 제거)
    if not os.path.exists(TEMPLATE_XLSX):
        raise FileNotFoundError(f"양식 템플릿이 없습니다: {TEMPLATE_XLSX} (legend.png 와 함께 스크립트 폴더에 두세요)")
    tplwb=openpyxl.load_workbook(TEMPLATE_XLSX)
    tpl=tplwb[tplwb.sheetnames[0]]
    nb=len(buckets); C_J=10; C_T=9+nb; GAP=(C_T+1,C_T+2,C_T+3)
    C_TOT=C_T+4; C_ISS=C_TOT+1; C_ZBEM=C_TOT+2; C_ADI=C_TOT+3; NCOL=C_ADI
    # 월 라벨(원본 형식: 'MARCH'26')
    MONTHS=["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE","JULY","AUGUST",
            "SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
    month_lbl=f"{MONTHS[today.month-1]}'{today.strftime('%y')}"

    def hdr_merges():
        # 헤더(1~5행) 병합을 템플릿에서 그대로 가져옴
        return [str(m) for m in tpl.merged_cells.ranges if m.min_row<=5]

    wb=Workbook(); first=True
    for sheet_name,_fam in SHEETS:
        rows=data_by_sheet.get(sheet_name,[])
        ws=wb.active if first else wb.create_sheet(); first=False
        ws.title=sheet_name+"".join(today_str.split("-")[1:3])
        # 열너비/숨김 복사
        for c in range(1,NCOL+1):
            cl=get_column_letter(c)
            if cl in tpl.column_dimensions:
                ws.column_dimensions[cl].width=tpl.column_dimensions[cl].width
        # 행높이(헤더) 복사
        for r in T_HDR:
            if r in tpl.row_dimensions: ws.row_dimensions[r].height=tpl.row_dimensions[r].height

        # ---- 헤더 1~5행: 스타일+값 그대로 복사 ----
        for r in T_HDR:
            for c in range(1,NCOL+1):
                s=tpl.cell(r,c); d=ws.cell(r,c)
                _copy_style(s,d)
                if s.value is not None: d.value=s.value
        # 병합 복사
        for mr in hdr_merges():
            try: ws.merge_cells(mr)
            except Exception: pass
        # 날짜 의존 값 덮어쓰기 (스타일은 유지)
        ws.cell(1,C_TOT).value=today                 # X1: 날짜(표시형식은 템플릿=인니 long date)
        ws.cell(3,C_J).value=month_lbl               # J3: 월 라벨
        for j,(label,_d,daynum,k) in enumerate(buckets):
            ws.cell(4,C_J+j).value=f"{_d[4:6]}-{_d[6:8]}"   # 4행: 날짜(MM-DD, 종합 동일)
            ws.cell(5,C_J+j).value=label             # 5행: D오프셋(색은 템플릿 그대로)

        # ---- 색상 범례 이미지 ----
        try:
            from openpyxl.drawing.image import Image as XLImage
            if os.path.exists(LEGEND_PNG):
                im=XLImage(LEGEND_PNG); im.width=130; im.height=30; ws.add_image(im,"E1")
        except Exception:
            pass

        # ---- 데이터 ----
        rr=6; grand=[0]*nb; gtot=0; run_start=rr; run_dong=None; pmerges=[]
        for row in rows:
            donor = T_DATA  # 대표 데이터행 스타일(테두리·폰트·정렬·표시형식)
            for c in range(1,NCOL+1):
                _copy_style(tpl.cell(donor,c), ws.cell(rr,c)); ws.cell(rr,c).fill=NOFILL
            ws.cell(rr,1,row["dong"]); ws.cell(rr,2,row["ic"]); ws.cell(rr,3,row["line"])
            ws.cell(rr,4,row["model"]); ws.cell(rr,5,row["style"])
            ws.cell(rr,6,row["color"] or None); ws.cell(rr,7,row["spray"] or None); ws.cell(rr,8,row["pad"] or None)
            ws.cell(rr,9,row["gen"])
            for j,(_l,_d,_n,k) in enumerate(buckets):
                v=row["cells"][j]; ws.cell(rr,C_J+j, v if v else None); grand[j]+=v
            ws.cell(rr,C_TOT,row["total"])   # 수식 대신 계산값(모든 뷰어에서 숫자 표시)
            ws.cell(rr,C_ISS,row["issue"] or None)
            ws.cell(rr,C_ZBEM,row["scan_bem"] or None); ws.cell(rr,C_ADI,row["scan_di"] or None)
            gtot+=row["total"]
            if row["dong"]!=run_dong:
                if run_dong is not None and rr-1>=run_start: pmerges.append((run_start,rr-1))
                run_dong=row["dong"]; run_start=rr
            rr+=1
        if run_dong is not None and rr-1>=run_start: pmerges.append((run_start,rr-1))
        # 마지막 데이터행은 하단 medium(표 박스 닫힘) 위해 row70 스타일로 덮기
        if rr>6:
            for c in range(1,NCOL+1):
                _copy_style(tpl.cell(T_LASTDATA,c), ws.cell(rr-1,c)); ws.cell(rr-1,c).fill=NOFILL
        # PLANT(동) 세로 병합
        for s,e in pmerges:
            if e>s:
                ws.merge_cells(start_row=s,start_column=1,end_row=e,end_column=1)

        # ---- GRAND TOTAL (템플릿 71행 스타일) ----
        for c in range(1,NCOL+1):
            _copy_style(tpl.cell(T_GTOT,c), ws.cell(rr,c))
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=9)
        ws.cell(rr,1,"GRAND TOTAL")
        for j,g in enumerate(grand): ws.cell(rr,C_J+j, g or None)
        ws.cell(rr,C_TOT,gtot)

        ws.sheet_view.showGridLines=False
        ws.freeze_panes=ws.cell(6,C_J)
        # ---- 페이지 설정(원본과 동일: landscape·A4·한 페이지 맞춤) ----
        ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
        ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
        ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_margins.left=ws.page_margins.right=ws.page_margins.top=ws.page_margins.bottom=1.0
        ws.print_area=f"A1:{get_column_letter(NCOL)}{rr}"
    return wb
