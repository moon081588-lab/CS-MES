#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CKP Manual Report — Balance ... by size 생성기
================================================
검증된 shortage-screen 엔진(P_MSPD90000S_Q_V14, END_ROUTING='Y'+CLOSING_YN='N'+BOM 색상)의
tidy 결과(LINE,MODEL_NAME,STYLE_CD,ITEM_CLASS,MCS_COLOR,FA_DATE,DIV,SIZE_CD,QTY)를 받아
원본 "3-2. Balance ... by size" 시트 레이아웃(사이즈 컬럼 PIVOT + Total/G-Total)으로 xlsx 생성.

사용:  python balance_bysize.py OUT.xlsx  SHEETNAME=csvpath  [SHEETNAME=csvpath ...]
  예:  python balance_bysize.py CKP_Balance_bysize.xlsx \
         "IP Prod by size=ip_prod.csv" "IP Outgoing by size=ip_out.csv" "PH by size=ph.csv"
CSV = sqlcl run-sql 결과 텍스트 그대로(맨 끝 'N rows selected.' 등 잡줄은 무시).
"""
import sys, csv, io
from collections import OrderedDict, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3A5F"; ALT = "EAF1F8"; SUB = "FFF2CC"; GT = "D9E1F2"
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDRFONT = Font(bold=True, color="FFFFFF", size=9, name="Malgun Gothic")
CELL = Font(size=9, name="Malgun Gothic")
BOLD = Font(bold=True, size=9, name="Malgun Gothic")

FIXED = ["Line", "Model Name", "Style Code", "Item Class", "MCS Color", "FA Date", "Div", "Shortage"]


def _num(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def read_tidy(path):
    with open(path, encoding="utf-8") as f:
        txt = f.read()
    rows = []
    for d in csv.DictReader(io.StringIO(txt)):
        d = {(k or "").strip().upper(): v for k, v in d.items()}
        q = _num(d.get("QTY"))
        if not q:
            continue
        rows.append({
            "line": (d.get("LINE") or "").strip(),
            "model": (d.get("MODEL_NAME") or "").strip(),
            "style": (d.get("STYLE_CD") or "").strip(),
            "ic": (d.get("ITEM_CLASS") or "").strip(),
            "color": (d.get("MCS_COLOR") or "").strip(),
            "fa": str(d.get("FA_DATE") or "").strip(),
            "div": (d.get("DIV") or "").strip(),
            "size": (d.get("SIZE_CD") or "").strip(),
            "qty": q,
        })
    return rows


def _size_sort_key(s):
    # 숫자 먼저, 그 뒤 T(하프) — '1' < '1T' < '2' ...
    import re
    m = re.match(r"(\d+)", s)
    n = int(m.group(1)) if m else 999
    return (n, 0 if s == (m.group(1) if m else s) else 1, s)


def build_sheet(ws, rows):
    sizes = sorted({r["size"] for r in rows if r["size"]}, key=_size_sort_key)
    headers = FIXED + sizes
    # 헤더
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = HDRFONT; cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    # row-key 그룹핑 (사이즈 pivot)
    key = lambda r: (r["line"], r["model"], r["style"], r["ic"], r["color"], r["fa"], r["div"])
    grouped = OrderedDict()
    for r in rows:
        grouped.setdefault(key(r), defaultdict(int))
        grouped[key(r)][r["size"]] += r["qty"]
    # 정렬: line, style, ic, fa
    ordered = sorted(grouped.items(), key=lambda kv: (kv[0][0], kv[0][2], kv[0][3], kv[0][5]))

    def size_cells(szmap):
        return [szmap.get(s, 0) for s in sizes]

    gtot_sizes = defaultdict(int); gtot_short = 0
    # G-Total 행(맨 위, 값은 나중에 채움) placeholder 위치 확보
    gt_row = 2
    ws.cell(gt_row, 1, "G-Total").font = BOLD
    for c in range(1, len(headers) + 1):
        ws.cell(gt_row, c).fill = PatternFill("solid", fgColor=GT)
        ws.cell(gt_row, c).border = BORDER
        ws.cell(gt_row, c).font = BOLD

    rr = 3
    cur_ls = None; ls_short = 0; ls_sizes = defaultdict(int); ls_start = None; alt = 0
    def flush_ls(endrow):
        nonlocal rr
        if cur_ls is None:
            return
        # Total 행
        ws.cell(rr, 7, "Total").font = BOLD
        ws.cell(rr, 8, ls_short).font = BOLD
        for i, s in enumerate(sizes):
            ws.cell(rr, 9 + i, ls_sizes.get(s, 0) or None).font = BOLD
        for c in range(1, len(headers) + 1):
            ws.cell(rr, c).fill = PatternFill("solid", fgColor=SUB)
            ws.cell(rr, c).border = BORDER

    for (k, szmap) in ordered:
        line, model, style, ic, color, fa, div = k
        short = sum(szmap.values())
        ls = (line, style)
        if cur_ls is not None and ls != cur_ls:
            flush_ls(rr); rr += 1
            cur_ls = None; ls_short = 0; ls_sizes = defaultdict(int)
        if cur_ls is None:
            cur_ls = ls; ls_short = 0; ls_sizes = defaultdict(int); alt ^= 1
        vals = [line, model, style, ic, color, fa, div, short] + size_cells(szmap)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rr, c, (v if (not isinstance(v, int) or v != 0) else None))
            cell.font = CELL; cell.border = BORDER
            if alt:
                cell.fill = PatternFill("solid", fgColor=ALT)
        ls_short += short
        for s in sizes:
            ls_sizes[s] += szmap.get(s, 0)
            gtot_sizes[s] += szmap.get(s, 0)
        gtot_short += short
        rr += 1
    flush_ls(rr); rr += 1
    # G-Total 값 채우기
    ws.cell(gt_row, 8, gtot_short).font = BOLD
    for i, s in enumerate(sizes):
        ws.cell(gt_row, 9 + i, gtot_sizes.get(s, 0) or None).font = BOLD
    # 열너비
    widths = [8, 24, 14, 10, 18, 11, 11, 10] + [6] * len(sizes)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"; ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    return gtot_short


def main():
    out = sys.argv[1]
    specs = sys.argv[2:]
    wb = openpyxl.Workbook(); first = True
    summary = []
    for spec in specs:
        name, path = spec.split("=", 1)
        rows = read_tidy(path)
        ws = wb.active if first else wb.create_sheet(); first = False
        ws.title = name[:31]
        gt = build_sheet(ws, rows)
        summary.append(f"· {name}: {len(rows)}행 / Shortage {gt:,}")
    wb.save(out)
    print("SAVED:", out)
    print("\n".join(summary))


if __name__ == "__main__":
    main()
