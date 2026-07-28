#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Changshin GMES — BALANCE OUTGOING 일일 자동발송 (독립 실행형)
================================================================
밑창 미드솔(IP·PH)+아웃솔(OS) '출고 부족분'을 라이브 DB에서 조회하여
원본 'BALANCE OUTGOING MARKET' 엑셀 양식 **그대로**(컬럼 A~I / 일자열 J~T 작업일 / 공백 U~W /
X=TOTAL · Y=ISSUE · Z=SCAN BEM/BEP · AA=SCAN DI CKP, 동(棟) 세로병합, GRAND TOTAL) 생성하여
사내 SMTP로 발송한다. Claude/계정 무관 순수 Python. OS 스케줄러로 매일 08:00 실행.

[채움 상태] COLOR = OCI.MSPD_BATCH_PLAN.MCS_COLOR_CD, SCAN BEM/BEP = OCI.POP_PCARD_SCAN(op BEM/BEP)
로 자동 채워진다. IP SPRAY(BEM) / PAD PRINTING(BEP) / SCAN DI CKP 는 소스(조인키) 미확인으로 아직
빈칸이며, 확인되면 color_specs()/scan_di_ckp() 한 곳만 채우면 자동 반영된다.

필요 패키지:  pip install oracledb openpyxl
사용:  python balance_outgoing_mailer.py [--dry-run | --test-db | --config 경로.ini]
"""
import os, sys, ssl, re, smtplib, argparse, configparser, logging, datetime, getpass, time, platform, subprocess
from email.message import EmailMessage

HERE = os.path.dirname(os.path.abspath(__file__))

# Windows 에서 stdout 이 파일/파이프면 ANSI 코드페이지로 인코딩되어 한글·기호 출력이
# UnicodeEncodeError 로 죽는다. 진입점에서 한 번 UTF-8 로 고정한다.
for _s in (sys.stdout, sys.stderr):
    try: _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception: pass

VERSION = "1.1"

def retry(fn, tries=3, delay=5, label="작업"):
    """일시 오류(DB/SMTP) 대비 재시도. 마지막 시도까지 실패하면 예외 재발생."""
    last=None
    for i in range(1,tries+1):
        try:
            return fn()
        except Exception as e:
            last=e
            if i<tries:
                LOG.warning(f"{label} 실패({i}/{tries}): {e} → {delay}초 후 재시도")
                time.sleep(delay)
    raise last

def load_config(path):
    cp = configparser.ConfigParser(interpolation=None)   # 비밀번호의 % 등 특수문자 안전
    if not cp.read(path, encoding="utf-8"): sys.exit(f"[설정오류] config 없음: {path}")
    # 값 뒤에 붙은 인라인 주석(;) 자동 제거 → 'window_before = 3   ; 설명' 같은 입력도 OK.
    # 단 비밀번호 계열은 값에 ;가 포함될 수 있어 건드리지 않음.
    KEEP = {"password", "wallet_password"}
    for sec in cp.sections():
        for k, v in cp[sec].items():
            if k not in KEEP and v and ";" in v:
                cp[sec][k] = v.split(";", 1)[0].strip()
    return cp

def save_password(path, section, value, key="password"):
    """config.ini의 [section] key 에 값을 저장하고 파일권한 600."""
    cp=configparser.ConfigParser(interpolation=None); cp.read(path,encoding="utf-8")
    if section not in cp: cp[section]={}
    cp[section][key]=value
    with open(path,"w",encoding="utf-8") as f: cp.write(f)
    try: os.chmod(path,0o600)
    except Exception: pass

def ensure_password(cfg, path, section, label):
    """비밀번호가 config에 있으면 그대로, 없으면 1회 숨김 입력받아 저장 후 반환."""
    v=cfg.get(section,"password",fallback="").strip()
    if v: return v
    if not sys.stdin.isatty():
        sys.exit(f"[설정] [{section}] password 가 비어 있습니다. 먼저 대화형으로 1회 실행해 저장하세요"
                 f" (예: python3 {os.path.basename(__file__)} --setup).")
    v=getpass.getpass(f"Enter {label} (saved once, hidden): ").strip()
    if not v: sys.exit("[설정] 비밀번호가 비어 있어 중단합니다.")
    if section not in cfg: cfg[section]={}
    cfg[section]["password"]=v
    save_password(path, section, v)
    LOG.info(f"[{section}] 비밀번호를 {path} 에 저장했습니다 (다음부터 묻지 않음).")
    return v

def setup_log():
    log=logging.getLogger("bo"); log.setLevel(logging.INFO)
    fmt=logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    sh=logging.StreamHandler(); sh.setFormatter(fmt); log.addHandler(sh)
    try:
        fh=logging.FileHandler(os.path.join(HERE,"balance_outgoing.log"),encoding="utf-8"); fh.setFormatter(fmt); log.addHandler(fh)
    except Exception: pass
    return log
LOG=setup_log()

SHEETS=[("IP",["II","IP"]),("PH",["PH","PP","CP"]),("OS",["OS"])]

# ============================================ 추후 연결 지점 (현재 DB에 데이터 없음 → 빈칸)
def color_specs(plant_cd, style_cd, item_class):
    """COLOR / IP SPRAY(BEM) / PAD PRINTING(BEP). [TODO] 색상 스펙 소스 확보 시 연결."""
    return ("", "", "")
def scan_di_ckp(plant_cd, fa_wc_cd, style_cd):
    """SCAN DI CKP. [TODO] DI 체크포인트 소스 확보 시 연결."""
    return ""
# ============================================

def build_buckets(today, before, after):
    """원본과 동일: 작업일(일요일 제외) 기준 D+before … DD … D-after.  반환 (label, yyyymmdd, 일, k)."""
    def walk(n, direction):
        days=[]; d=today
        while len(days)<n:
            d=d+datetime.timedelta(days=direction)
            if d.weekday()!=6: days.append(d)        # 일요일 제외(토요일은 작업일)
        return days
    seq=list(reversed(walk(before,-1)))+[today]+walk(after,+1)
    labels=[f"D+{k}" for k in range(before,0,-1)]+["DD"]+[f"D-{k}" for k in range(1,after+1)]
    ks=list(range(before,0,-1))+[0]+[-k for k in range(1,after+1)]
    return [(labels[i],seq[i].strftime("%Y%m%d"),seq[i].day,ks[i]) for i in range(len(seq))]

def dong(wcg):
    m=re.match(r"[A-Za-z]+",(wcg or "").strip())
    return m.group(0) if m else (wcg or "").strip()

def report_dir(cfg):
    """생성된 리포트 저장 폴더. 기본 = 스크립트 상위의 report/ (OneDrive 공유 폴더)."""
    d=cfg.get("report","output_dir",fallback="").strip()
    d=os.path.abspath(d) if d else os.path.abspath(os.path.join(HERE,"..","report"))
    os.makedirs(d,exist_ok=True)
    return d

def site_today(cfg=None):
    """현장(공장) 기준 '오늘'. 실행 PC 의 로컬 날짜(datetime.date.today())를 쓰면 안 된다.

    같은 순간에 세 개의 날짜가 존재한다 (2026-07-28 실측):
      DB(SYSDATE, UTC)  00:19 / 한국 PC(KST)  09:19 / 현장 CKP(WIB, UTC+7)  07:19
    한국에서 돌리든 현지에서 돌리든 리포트 기준일이 같아야 하므로 현장 타임존으로 계산한다.
    우선순위: config [report] site_timezone > 환경변수 CSMES_TZ > Asia/Jakarta.
    (Windows 에서 zoneinfo 를 쓰려면 `pip install tzdata` 필요 — 없으면 로컬 날짜로 폴백.)
    """
    tz = ""
    try:
        if cfg is not None:
            tz = cfg.get("report", "site_timezone", fallback="").strip()
        else:
            c = configparser.ConfigParser(interpolation=None, inline_comment_prefixes=(";",))
            c.read(os.path.join(HERE, "config.ini"), encoding="utf-8")
            tz = c.get("report", "site_timezone", fallback="").strip()
    except Exception:
        tz = ""
    tz = tz or os.environ.get("CSMES_TZ", "").strip() or "Asia/Jakarta"
    try:
        from zoneinfo import ZoneInfo
        return datetime.datetime.now(ZoneInfo(tz)).date()
    except Exception as e:
        LOG.warning(f"타임존 {tz} 을 쓸 수 없어 실행 PC 로컬 날짜를 씁니다"
                    f"(Windows 는 pip install tzdata 필요): {e}")
        return datetime.date.today()


def wallet_dir(cfg):
    """월렛 폴더. config 값이 비어 있으면 스크립트와 같은 폴더의 wallet/ 을 쓴다.
    (절대경로를 config 에 박지 않아야 폴더를 옮기거나 다른 PC 로 배포해도 그대로 동작한다.)"""
    d=cfg.get("db","wallet_dir",fallback="").strip()
    if not d:
        cand=os.path.join(HERE,"wallet")
        d=cand if os.path.isdir(cand) else ""
    return os.path.abspath(d) if d else ""

def sync_sqlnet_ora(wdir):
    """wallet/sqlnet.ora 의 WALLET_LOCATION DIRECTORY 를 현재 실제 경로로 교정.
    다른 PC(또는 다른 폴더)에서 복사해 온 월렛은 남의 절대경로가 박혀 있어
    ORA-12154 / ORA-28759 / ORA-29024 를 일으킨다. 매 실행마다 조용히 맞춰준다."""
    p=os.path.join(wdir,"sqlnet.ora")
    if not (wdir and os.path.isfile(p)): return
    try:
        s=open(p,encoding="utf-8").read()
        new=re.sub(r'DIRECTORY\s*=\s*"[^"]*"', 'DIRECTORY="%s"' % wdir.replace("\\","\\\\"), s)
        if new!=s:
            open(p,"w",encoding="utf-8").write(new)
            LOG.info(f"sqlnet.ora WALLET_LOCATION 교정 → {wdir}")
    except Exception as e:
        LOG.warning(f"sqlnet.ora 교정 실패(무시하고 진행): {e}")

def _ic_hint():
    """Instant Client 설치 안내 — OS 별로 다르게."""
    s=platform.system()
    if s=="Windows":
        return ("Windows: https://www.oracle.com/database/technology/instant-client-downloads.html 에서 "
                "'Basic Package (ZIP, x64)' 를 받아 C:\\oracle\\instantclient 에 풀고, "
                "config.ini [db] oracle_client_lib 에 그 폴더 경로를 적으세요.")
    if s=="Darwin":
        return "macOS: brew tap InstantClientTap/instantclient && brew install instantclient-basic (csmes.sh 가 자동 시도)"
    return "Linux: oracle-instantclient-basic 패키지 설치 후 config.ini [db] oracle_client_lib 지정"

def db_connect(cfg):
    """접속 모드 3 종.
      thick — Instant Client + cwallet.sso 자동로그인 (월렛 비번 불필요)
      thin  — Instant Client 불필요, 대신 월렛 비번(PEM pass phrase) 필요
      auto  — thick 먼저 시도, 실패하면 thin 으로 폴백 (기본값·배포 권장)
    """
    import oracledb
    mode=cfg.get("db","mode",fallback="auto").strip().lower()
    if mode not in ("thick","thin","auto"): mode="auto"
    wdir=wallet_dir(cfg); wpw=cfg.get("db","wallet_password",fallback="").strip()
    sync_sqlnet_ora(wdir)
    base={"user":cfg.get("db","user"),"password":cfg.get("db","password"),"dsn":cfg.get("db","dsn")}

    def _try_thick():
        lib=cfg.get("db","oracle_client_lib",fallback="").strip() or None
        try:
            oracledb.init_oracle_client(lib_dir=lib)
        except Exception as e:
            if "already" not in str(e).lower():
                raise RuntimeError(f"Oracle Instant Client 를 찾지 못했습니다(thick). {_ic_hint()} / 원오류: {e}")
        kw=dict(base)
        if wdir: kw["config_dir"]=wdir      # sqlnet.ora 의 WALLET_LOCATION 으로 자동로그인
        LOG.info(f"DB 접속 시도: mode=thick dsn={kw['dsn']} wallet={'Y' if wdir else 'N'}")
        return oracledb.connect(**kw)

    def _try_thin():
        kw=dict(base)
        if wdir:
            kw["config_dir"]=wdir; kw["wallet_location"]=wdir
            if wpw: kw["wallet_password"]=wpw
            elif os.path.exists(os.path.join(wdir,"ewallet.pem")):
                raise RuntimeError(
                    "[thin 모드] 월렛 비밀번호(PEM pass phrase)가 필요합니다.\n"
                    "  · config.ini [db] wallet_password 에 넣거나,\n"
                    f"  · mode=thick 로 바꾸고 Instant Client 를 설치하세요. {_ic_hint()}")
        LOG.info(f"DB 접속 시도: mode=thin dsn={kw['dsn']} wallet={'Y' if wdir else 'N'}")
        return oracledb.connect(**kw)

    if mode=="thick": return _try_thick()
    if mode=="thin":  return _try_thin()
    try:
        return _try_thick()
    except Exception as e:
        LOG.warning(f"thick 실패 → thin 으로 재시도: {str(e)[:160]}")
        try:
            return _try_thin()
        except Exception as e2:
            raise RuntimeError(
                "DB 접속 실패 (thick·thin 모두).\n"
                f"  · thick: {str(e)[:200]}\n"
                f"  · thin : {str(e2)[:200]}\n"
                f"  · 월렛 폴더: {wdir or '(미설정)'}\n"
                f"  · {_ic_hint()}")

_COLOR_JOIN = (   # style_cd 별 대표색(NONE 제외 최빈) — BATCH_PLAN.MCS_COLOR_CD
    "LEFT JOIN (SELECT style_cd, mcs_color_cd FROM ("
    "  SELECT style_cd, mcs_color_cd, ROW_NUMBER() OVER (PARTITION BY style_cd"
    "    ORDER BY CASE WHEN mcs_color_cd='NONE' THEN 1 ELSE 0 END, COUNT(*) DESC) rn"
    "  FROM OCI.MSPD_BATCH_PLAN WHERE mcs_color_cd IS NOT NULL GROUP BY style_cd, mcs_color_cd"
    ") WHERE rn=1) cc ON cc.style_cd=o.style_cd"
)

def fetch_sheet(conn, families, plants, d_from, d_to, strict=True):
    """미출고 부족분 조회.
    strict=True  : GMES 정식 프로시저 P_MSPD90000S_Q_V14 'O'(Outgoing) 분기와 동일 로직
                   (엄격 CLOSING_YN='N' + '다른 창고로 나가는 MOVE 존재' EXISTS). OCI 동기화 전제.
    strict=False : OCI 미동기화 임시용(느슨한 마감, EXISTS 생략). 데이터가 덜 동기화돼도 산출.
    반환: (wcg, plant, item_class, fa_wc, model, gen, style, fa_date, qty, color)
    """
    fam=",".join(f":f{i}" for i in range(len(families))); pl=",".join(f":p{i}" for i in range(len(plants)))
    if strict:
        sql=f"""
          SELECT NVL(w2.wc_group_cd,' ') wcg, o.plant_cd, o.item_class, o.fa_wc_cd,
                 NVL(i.model_name,' ') model, NVL(i.gender,' ') gen, o.style_cd, o.fa_date,
                 SUM(o.out_qty) qty, MAX(cc.mcs_color_cd) color
          FROM (
            SELECT R.FA_WC_CD, R.ITEM_CLASS, R.FA_DATE, R.STYLE_CD, R.PLANT_CD,
                   R.PLAN_PROD_WC_CD, R.PROD_GROUP_NO, SUM(R.PCARD_QTY) out_qty
            FROM OCI.MSPD_PCARD_RESULT R
            WHERE R.FA_DATE BETWEEN :d_from AND :d_to
              AND R.PLANT_CD IN ({pl}) AND R.PROD_MOVE_TYPE='PROD'
              AND R.ITEM_CLASS_TYPE IN ({fam})
              AND (R.PROD_GROUP_NO, R.PLANT_CD) IN (SELECT PROD_GROUP_NO, PLANT_CD FROM OCI.MSPD_PROD_GROUP WHERE CLOSING_YN='N')
              AND R.END_ROUTING_YN='Y' AND R.OUT_DATE='19991231'
            GROUP BY R.FA_WC_CD,R.ITEM_CLASS,R.FA_DATE,R.STYLE_CD,R.PLANT_CD,R.PLAN_PROD_WC_CD,R.PROD_GROUP_NO
            HAVING SUM(R.PCARD_QTY)>0
          ) o
          JOIN OCI.MSBS_WORK_CENTER w  ON o.plant_cd=w.plant_cd AND o.plan_prod_wc_cd=w.wc_cd
          LEFT JOIN OCI.MSBS_ITEM_STYLE i ON i.style_cd=o.style_cd
          LEFT JOIN OCI.MSBS_WORK_CENTER w2 ON w2.plant_cd=o.plant_cd AND w2.wc_cd=o.fa_wc_cd
          {_COLOR_JOIN}
          WHERE EXISTS (SELECT 1 FROM OCI.MSPD_PCARD_RESULT I
                          JOIN OCI.MSBS_WORK_CENTER WC ON I.PLANT_CD=WC.PLANT_CD AND I.PLAN_PROD_WC_CD=WC.WC_CD
                         WHERE I.PROD_GROUP_NO=o.PROD_GROUP_NO AND I.ITEM_CLASS=o.ITEM_CLASS
                           AND I.PROD_MOVE_TYPE='MOVE' AND I.END_ROUTING_YN='Y'
                           AND WC.BASE_WH_CD<>w.BASE_WH_CD)
          GROUP BY NVL(w2.wc_group_cd,' '), o.plant_cd, o.item_class, o.fa_wc_cd,
                   NVL(i.model_name,' '), NVL(i.gender,' '), o.style_cd, o.fa_date
        """
    else:
        sql=f"""
          SELECT NVL(w.wc_group_cd,' ') wcg, o.plant_cd, o.item_class, o.fa_wc_cd,
                 NVL(i.model_name,' ') model, NVL(i.gender,' ') gen, o.style_cd, o.fa_date,
                 SUM(o.out_qty) qty, MAX(cc.mcs_color_cd) color
          FROM (
            SELECT R.FA_WC_CD, R.ITEM_CLASS, R.FA_DATE, R.STYLE_CD, R.PLANT_CD, R.PROD_GROUP_NO,
                   SUM(R.PCARD_QTY) out_qty
            FROM OCI.MSPD_PCARD_RESULT R
            WHERE R.FA_DATE BETWEEN :d_from AND :d_to
              AND R.PLANT_CD IN ({pl}) AND R.PROD_MOVE_TYPE='PROD'
              AND R.ITEM_CLASS_TYPE IN ({fam})
              AND R.END_ROUTING_YN='Y' AND R.OUT_DATE='19991231'
              AND NOT EXISTS (SELECT 1 FROM OCI.MSPD_PROD_GROUP g
                               WHERE g.prod_group_no=R.prod_group_no AND g.plant_cd=R.plant_cd AND g.closing_yn='Y')
            GROUP BY R.FA_WC_CD,R.ITEM_CLASS,R.FA_DATE,R.STYLE_CD,R.PLANT_CD,R.PROD_GROUP_NO
            HAVING SUM(R.PCARD_QTY)>0
          ) o
          LEFT JOIN OCI.MSBS_ITEM_STYLE i ON i.style_cd=o.style_cd
          LEFT JOIN OCI.MSBS_WORK_CENTER w ON w.plant_cd=o.plant_cd AND w.wc_cd=o.fa_wc_cd
          {_COLOR_JOIN}
          GROUP BY NVL(w.wc_group_cd,' '), o.plant_cd, o.item_class, o.fa_wc_cd,
                   NVL(i.model_name,' '), NVL(i.gender,' '), o.style_cd, o.fa_date
        """
    b={f"f{i}":v for i,v in enumerate(families)}; b.update({f"p{i}":v for i,v in enumerate(plants)})
    b.update({"d_from":d_from,"d_to":d_to})
    cur=conn.cursor(); cur.execute(sql,b); rows=cur.fetchall(); cur.close(); return rows

def fetch_scan_bembep(conn, plants, d_from, d_to):
    pl=",".join(f":p{i}" for i in range(len(plants)))
    sql=f"""SELECT plant_cd, fa_wc_cd, style_cd, SUM(prod_qty) q FROM OCI.POP_PCARD_SCAN
            WHERE op_cd IN ('BEM','BEP') AND NVL(cancel_flag,'N')<>'Y' AND plant_cd IN ({pl})
              AND scan_ymd BETWEEN :d_from AND :d_to
            GROUP BY plant_cd, fa_wc_cd, style_cd"""
    b={f"p{i}":v for i,v in enumerate(plants)}; b.update({"d_from":d_from,"d_to":d_to})
    cur=conn.cursor(); cur.execute(sql,b); d={(p,w,s):(q or 0) for p,w,s,q in cur.fetchall()}; cur.close(); return d

def _clean_color(raw):
    """'OBSIDIAN(45B)' → 'OBSIDIAN', 'NONE'·공란 → ''."""
    c=(raw or "").split("(")[0].strip()
    return "" if c.upper()=="NONE" else c

def pivot(rows, buckets, scan_map):
    bdates=[b[1] for b in buckets]; table={}; colors={}
    for row in rows:
        wcg,plant,ic,line,model,gen,style,fa,qty = row[:9]
        color = row[9] if len(row)>9 else None      # 10번째 컬럼=대표색(있으면)
        dng=dong(wcg) or plant
        key=(dng,plant,ic,line,(model or " ").strip(),(gen or " ").strip(),style)
        table.setdefault(key,{})[fa]=table.setdefault(key,{}).get(fa,0)+(qty or 0)
        if color and style not in colors: colors[style]=color
    out=[]
    for key,dd in table.items():
        cells=[dd.get(bd,0) for bd in bdates]; total=sum(cells)
        if total<=0: continue
        dng,plant,ic,line,model,gen,style=key
        _c,spray,pad=color_specs(plant,style,ic)     # spray/pad 는 아직 빈칸(후크)
        color=_clean_color(colors.get(style))
        out.append({"dong":dng,"ic":ic,"line":line,"model":model,"style":style,"color":color,
                    "spray":spray,"pad":pad,"gen":gen,"cells":cells,"total":total,"issue":"",
                    "scan_bem":scan_map.get((plant,line,style),0),"scan_di":scan_di_ckp(plant,line,style)})
    out.sort(key=lambda r:(r["dong"],r["line"],r["style"]))
    return out

# ---------- Excel: 원본을 '스타일 템플릿'으로 사용해 100% 동일 양식 복제 ----------
# 원본(2.3 BALANCE OUTGOING)의 IP0306 시트를 report_template.xlsx(스타일 도너)로 보관하고,
# 각 셀의 폰트·채움·테두리·정렬·표시형식을 그대로 복사한다. 날짜의존 값(날짜/월/일/D오프셋)과
# 본문 데이터만 새로 써넣는다. 색상범례는 원본 이미지를 그대로 잘라낸 legend.png 를 삽입한다.
import os
from copy import copy as _copy
_HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_XLSX = os.path.join(_HERE, "report_template.xlsx")
LEGEND_PNG    = os.path.join(_HERE, "legend.png")

# 템플릿 기준 행: 헤더 1~5, 대표 데이터행 6, 마지막 데이터행 70, GRAND TOTAL 71
T_HDR=range(1,6); T_DATA=6; T_LASTDATA=70; T_GTOT=71; T_NCOL=27

def _copy_style(src, dst):
    if src.has_style:
        dst.font=_copy(src.font); dst.fill=_copy(src.fill)
        dst.border=_copy(src.border); dst.alignment=_copy(src.alignment)
        dst.number_format=src.number_format; dst.protection=_copy(src.protection)

def build_workbook(data_by_sheet, buckets, today, today_str):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    NOFILL=PatternFill(fill_type=None)   # 데이터행은 흰색(원본의 수작업 강조색 제거)
    if not os.path.exists(TEMPLATE_XLSX):
        raise FileNotFoundError(f"양식 템플릿이 없습니다: {TEMPLATE_XLSX} (legend.png 와 함께 스크립트 폴더에 두세요)")
    tplwb=openpyxl.load_workbook(TEMPLATE_XLSX)
    tpl=tplwb[tplwb.sheetnames[0]]
    nb=len(buckets); C_J=10; C_T=9+nb; GAP=(C_T+1,C_T+2,C_T+3)
    C_TOT=C_T+4; C_ISS=C_TOT+1; C_ZBEM=C_TOT+2; C_ADI=C_TOT+3; NCOL=C_ADI
    # 월 라벨(원본 형식: 'MARCH'26')
    MONTHS=["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE","JULY","AUGUST",
            "SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
    month_lbl=f"{MONTHS[today.month-1]}'{today.strftime('%y')}"

    def hdr_merges():
        # 헤더(1~5행) 병합을 템플릿에서 그대로 가져옴
        return [str(m) for m in tpl.merged_cells.ranges if m.min_row<=5]

    wb=Workbook(); first=True
    for sheet_name,_fam in SHEETS:
        rows=data_by_sheet.get(sheet_name,[])
        ws=wb.active if first else wb.create_sheet(); first=False
        ws.title=sheet_name+"".join(today_str.split("-")[1:3])
        # 열너비/숨김 복사
        for c in range(1,NCOL+1):
            cl=get_column_letter(c)
            if cl in tpl.column_dimensions:
                ws.column_dimensions[cl].width=tpl.column_dimensions[cl].width
        # 행높이(헤더) 복사
        for r in T_HDR:
            if r in tpl.row_dimensions: ws.row_dimensions[r].height=tpl.row_dimensions[r].height

        # ---- 헤더 1~5행: 스타일+값 그대로 복사 ----
        for r in T_HDR:
            for c in range(1,NCOL+1):
                s=tpl.cell(r,c); d=ws.cell(r,c)
                _copy_style(s,d)
                if s.value is not None: d.value=s.value
        # 병합 복사
        for mr in hdr_merges():
            try: ws.merge_cells(mr)
            except Exception: pass
        # 날짜 의존 값 덮어쓰기 (스타일은 유지)
        ws.cell(1,C_TOT).value=today                 # X1: 날짜(표시형식은 템플릿=인니 long date)
        ws.cell(3,C_J).value=month_lbl               # J3: 월 라벨
        for j,(label,_d,daynum,k) in enumerate(buckets):
            ws.cell(4,C_J+j).value=daynum            # 4행: 일
            ws.cell(5,C_J+j).value=label             # 5행: D오프셋(색은 템플릿 그대로)

        # ---- 색상 범례 이미지 ----
        try:
            from openpyxl.drawing.image import Image as XLImage
            if os.path.exists(LEGEND_PNG):
                im=XLImage(LEGEND_PNG); im.width=130; im.height=30; ws.add_image(im,"E1")
        except Exception:
            pass

        # ---- 데이터 ----
        rr=6; grand=[0]*nb; gtot=0; run_start=rr; run_dong=None; pmerges=[]
        for row in rows:
            donor = T_DATA  # 대표 데이터행 스타일(테두리·폰트·정렬·표시형식)
            for c in range(1,NCOL+1):
                _copy_style(tpl.cell(donor,c), ws.cell(rr,c)); ws.cell(rr,c).fill=NOFILL
            ws.cell(rr,1,row["dong"]); ws.cell(rr,2,row["ic"]); ws.cell(rr,3,row["line"])
            ws.cell(rr,4,row["model"]); ws.cell(rr,5,row["style"])
            ws.cell(rr,6,row["color"] or None); ws.cell(rr,7,row["spray"] or None); ws.cell(rr,8,row["pad"] or None)
            ws.cell(rr,9,row["gen"])
            for j,(_l,_d,_n,k) in enumerate(buckets):
                v=row["cells"][j]; ws.cell(rr,C_J+j, v if v else None); grand[j]+=v
            ws.cell(rr,C_TOT,row["total"])   # 수식 대신 계산값(모든 뷰어에서 숫자 표시)
            ws.cell(rr,C_ISS,row["issue"] or None)
            ws.cell(rr,C_ZBEM,row["scan_bem"] or None); ws.cell(rr,C_ADI,row["scan_di"] or None)
            gtot+=row["total"]
            if row["dong"]!=run_dong:
                if run_dong is not None and rr-1>=run_start: pmerges.append((run_start,rr-1))
                run_dong=row["dong"]; run_start=rr
            rr+=1
        if run_dong is not None and rr-1>=run_start: pmerges.append((run_start,rr-1))
        # 마지막 데이터행은 하단 medium(표 박스 닫힘) 위해 row70 스타일로 덮기
        if rr>6:
            for c in range(1,NCOL+1):
                _copy_style(tpl.cell(T_LASTDATA,c), ws.cell(rr-1,c)); ws.cell(rr-1,c).fill=NOFILL
        # PLANT(동) 세로 병합
        for s,e in pmerges:
            if e>s:
                ws.merge_cells(start_row=s,start_column=1,end_row=e,end_column=1)

        # ---- GRAND TOTAL (템플릿 71행 스타일) ----
        for c in range(1,NCOL+1):
            _copy_style(tpl.cell(T_GTOT,c), ws.cell(rr,c))
        ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=9)
        ws.cell(rr,1,"GRAND TOTAL")
        for j,g in enumerate(grand): ws.cell(rr,C_J+j, g or None)
        ws.cell(rr,C_TOT,gtot)

        ws.sheet_view.showGridLines=False
        ws.freeze_panes=ws.cell(6,C_J)
        # ---- 페이지 설정(원본과 동일: landscape·A4·한 페이지 맞춤) ----
        ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
        ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
        ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_margins.left=ws.page_margins.right=ws.page_margins.top=ws.page_margins.bottom=1.0
        ws.print_area=f"A1:{get_column_letter(NCOL)}{rr}"
    return wb

def send_mail(cfg, xlsx_path, summary, today_str):
    host=cfg.get("smtp","host"); port=cfg.getint("smtp","port",fallback=587)
    user=cfg.get("smtp","user",fallback="").strip(); pw=cfg.get("smtp","password",fallback="").strip()
    use_tls=cfg.getboolean("smtp","use_tls",fallback=True); sender=cfg.get("smtp","from")
    recips=[x.strip() for x in cfg.get("report","recipients").split(",") if x.strip()]
    if not recips: sys.exit("[설정오류] recipients 비어 있음")
    msg=EmailMessage(); msg["Subject"]=f"[BALANCE OUTGOING] {today_str} 밑창 출고 부족분 (미드솔·아웃솔)"
    msg["From"]=sender; msg["To"]=", ".join(recips)
    link=cfg.get("report","share_link",fallback="").strip()
    link_line=f"\n📎 리포트 폴더(엑셀 다운로드): {link}\n" if link else ""
    msg.set_content("안녕하십니까.\n\n"
        f"{today_str} 기준 밑창 출고 부족분(BALANCE OUTGOING)을 첨부드립니다. (라이브 DB 자동생성)\n\n{summary}\n"
        f"{link_line}"
        "· 시트: IP(미드솔 IP사출) / PH(미드솔 파일론) / OS(아웃솔)\n"
        "· COLOR·SCAN BEM/BEP 는 자동 채움. IP SPRAY / PAD PRINTING / SCAN DI CKP 는 소스 확인 전이라 빈칸입니다.\n\n자동 발송 메일입니다.\n")
    with open(xlsx_path,"rb") as f:
        msg.add_attachment(f.read(),maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",filename=os.path.basename(xlsx_path))
    LOG.info(f"메일 발송: {host}:{port} → {recips}")
    if use_tls:
        with smtplib.SMTP(host,port,timeout=60) as s:
            s.starttls(context=ssl.create_default_context())
            if user: s.login(user,pw)
            s.send_message(msg)
    else:
        with smtplib.SMTP(host,port,timeout=60) as s:
            if user: s.login(user,pw)
            s.send_message(msg)
    LOG.info("메일 발송 완료")

def _smtp_send(cfg, msg):
    """[smtp] 설정으로 EmailMessage 발송 (공통)."""
    host=cfg.get("smtp","host"); port=cfg.getint("smtp","port",fallback=587)
    user=cfg.get("smtp","user",fallback="").strip(); pw=cfg.get("smtp","password",fallback="").strip()
    use_tls=cfg.getboolean("smtp","use_tls",fallback=True)
    LOG.info(f"SMTP 접속: {host}:{port} tls={use_tls} auth={'Y' if user else 'N'}")
    with smtplib.SMTP(host,port,timeout=60) as s:
        if use_tls: s.starttls(context=ssl.create_default_context())
        if user: s.login(user,pw)
        s.send_message(msg)

def send_test_mail(cfg):
    """DB 없이 SMTP 전송만 검증. 샘플 xlsx가 있으면 첨부."""
    sender=cfg.get("smtp","from")
    recips=[x.strip() for x in cfg.get("report","recipients").split(",") if x.strip()]
    if not recips: sys.exit("[설정오류] [report] recipients 비어 있음")
    now=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    msg=EmailMessage()
    msg["Subject"]=f"[TEST] BALANCE OUTGOING mailer SMTP 점검 {now}"
    msg["From"]=sender; msg["To"]=", ".join(recips)
    msg.set_content(f"SMTP 전송 테스트 메일입니다. (생성시각 {now})\n\n"
                    "이 메일이 도착했다면 메일 전송기 설정이 정상입니다.\n"
                    "— balance_outgoing_mailer.py --test-mail")
    # 같은 폴더에 샘플이 있으면 첨부(첨부 동작까지 확인)
    for fn in sorted(os.listdir(HERE)):
        if fn.startswith("SAMPLE_BALANCE_OUTGOING") and fn.endswith(".xlsx"):
            with open(os.path.join(HERE,fn),"rb") as f:
                msg.add_attachment(f.read(),maintype="application",
                    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",filename=fn)
            LOG.info(f"테스트 첨부: {fn}"); break
    LOG.info(f"테스트 메일 발송 → {recips}")
    _smtp_send(cfg,msg)
    LOG.info("테스트 메일 발송 완료")
    print(f"OK: 테스트 메일을 {', '.join(recips)} 로 보냈습니다. 받은편지함을 확인하세요.")

def build_body_summary(data):
    """메일 본문용 요약: 시트별 총잔량 + 전체 부족 TOP5."""
    lines=[]; grand=0; allrows=[]
    for name,_f in SHEETS:
        rows=data.get(name,[]); t=sum(r["total"] for r in rows); grand+=t
        lines.append(f"· {name}: {len(rows)}개 스타일 / 잔량 {t:,}족")
        for r in rows: allrows.append((name,r))
    lines.append(f"· 합계: 잔량 {grand:,}족")
    top=sorted(allrows,key=lambda x:x[1]['total'],reverse=True)[:5]
    if top:
        lines.append(""); lines.append("[부족 TOP5]")
        for i,(nm,r) in enumerate(top,1):
            lines.append(f"  {i}. [{nm}] {r.get('model','')} / {r.get('style','')} — {r['total']:,}족"
                         f" (동 {r.get('dong','')}, {r.get('line','')})")
    return "\n".join(lines)

def make_demo_data(buckets):
    """DB 없이 전체 흐름을 시연하기 위한 샘플 데이터(실제 양식 그대로)."""
    nb=len(buckets)
    def row(dong,ic,line,model,style,gen,vals,scan):
        cells=[0]*nb
        for idx,v in vals:
            if 0<=idx<nb: cells[idx]=v
        return {"dong":dong,"ic":ic,"line":line,"model":model,"style":style,"color":"","spray":"","pad":"",
                "gen":gen,"cells":cells,"total":sum(cells),"issue":"","scan_bem":scan,"scan_di":""}
    return {
      "IP":[row("A","IP01","FGA04","AIR MAX 2017 (M)","849559-405","ME",[(1,180)],0),
            row("A","II01","FGA05","AIR MAX 90 LTR (TD)","CD6868-100","TD",[(2,128)],0)],
      "PH":[row("B","PH01","FGB10","PEGASUS 41 (W)","FD2723-100","WO",[(3,640),(4,320)],500),
            row("B","CP02","FGB12","INVINCIBLE 3 (M)","DR2615-001","ME",[(0,210)],0)],
      "OS":[row("C","OS01","FGC07","METCON 9 (M)","IH7446-001","ME",[(3,1540)],1200),
            row("C","OS01","FGC08","VAPORFLY 3 (W)","FD6556-100","WO",[(5,300)],0)],
    }

def send_failure_mail(cfg, err_text, when_str):
    """파이프라인 실패 시 원인을 담아 best-effort 로 알림 발송."""
    try:
        sender=cfg.get("smtp","from",fallback="").strip()
        recips=[x.strip() for x in cfg.get("report","recipients",fallback="").split(",") if x.strip()]
        if not (sender and recips): LOG.error("실패 알림 생략(SMTP/수신자 미설정)"); return
        msg=EmailMessage()
        msg["Subject"]=f"[BALANCE OUTGOING 실패] {when_str} 리포트 작업 오류"
        msg["From"]=sender; msg["To"]=", ".join(recips)
        msg.set_content(f"{when_str} BALANCE OUTGOING 자동 작업이 실패했습니다.\n\n"
                        f"[원인]\n{err_text}\n\n로그: {os.path.join(HERE,'balance_outgoing.log')}\n"
                        "자동 발송된 실패 알림입니다.")
        _smtp_send(cfg,msg); LOG.info("실패 알림 메일 발송 완료")
    except Exception as e:
        LOG.error(f"실패 알림 메일도 실패: {e}")

def run_doctor(cfg, path):
    """라이브러리/템플릿/config/DB/SMTP 일괄 점검 → ✅/❌."""
    ok=True
    def line(name, good, detail=""):
        nonlocal ok
        if not good: ok=False
        print(("✅ " if good else "❌ ")+name+(f" — {detail}" if detail else ""))
    print(f"=== CS-MES doctor (v{VERSION}) ===")
    try: import oracledb; line("oracledb 설치", True, getattr(oracledb,'__version__',''))
    except Exception as e: line("oracledb 설치", False, str(e))
    try: import openpyxl; line("openpyxl 설치", True, getattr(openpyxl,'__version__',''))
    except Exception as e: line("openpyxl 설치", False, str(e))
    line("report_template.xlsx", os.path.exists(TEMPLATE_XLSX), TEMPLATE_XLSX)
    line("legend.png", os.path.exists(LEGEND_PNG), LEGEND_PNG)
    for sec,key in [("db","dsn"),("smtp","host"),("smtp","from"),("report","recipients"),("report","plants")]:
        v=cfg.get(sec,key,fallback="").strip(); line(f"config [{sec}] {key}", bool(v), v or "비어있음")
    for sec in ("db","smtp"):
        has=bool(cfg.get(sec,"password",fallback="").strip())
        line(f"config [{sec}] password", has, "저장됨" if has else "미저장 → --setup")
    try:
        conn=db_connect(cfg); cur=conn.cursor(); cur.execute("SELECT 1 FROM DUAL"); cur.fetchone(); conn.close()
        line("DB 접속(SELECT 1)", True)
    except Exception as e: line("DB 접속(SELECT 1)", False, str(e)[:120])
    try:
        host=cfg.get("smtp","host"); port=cfg.getint("smtp","port",fallback=587)
        user=cfg.get("smtp","user",fallback="").strip(); pw=cfg.get("smtp","password",fallback="").strip()
        use_tls=cfg.getboolean("smtp","use_tls",fallback=True)
        with smtplib.SMTP(host,port,timeout=30) as s:
            if use_tls: s.starttls(context=ssl.create_default_context())
            if user: s.login(user,pw)
        line("SMTP 접속/로그인", True)
    except Exception as e: line("SMTP 접속/로그인", False, str(e)[:120])
    print("="*32)
    print("전체 정상 ✅ — 'csmes' 로 실행하세요" if ok else "문제 발견 ❌ — 위 ❌ 항목을 확인하세요")
    return 0 if ok else 1

SCHED_LABEL="com.changshin.balanceoutgoing"
def _runner():
    """OS 별 실행 래퍼. Windows 는 csmes.sh(bash) 를 쓸 수 없으므로 csmes.bat → python 순."""
    if platform.system()=="Windows":
        bat=os.path.join(HERE,"csmes.bat")
        if os.path.exists(bat): return ("cmd.exe", bat)
        return (sys.executable, os.path.join(HERE,os.path.basename(__file__)))
    sh=os.path.join(HERE,"csmes.sh")
    return ("/bin/bash", sh) if os.path.exists(sh) else (sys.executable, os.path.join(HERE,os.path.basename(__file__)))

def install_schedule(hour=8, minute=0):
    sysname=platform.system(); prog,arg=_runner(); logp=os.path.join(HERE,"cron.log")
    if sysname=="Darwin":
        plist=os.path.expanduser(f"~/Library/LaunchAgents/{SCHED_LABEL}.plist")
        os.makedirs(os.path.dirname(plist),exist_ok=True)
        xml=f'''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0"><dict>
  <key>Label</key><string>{SCHED_LABEL}</string>
  <key>ProgramArguments</key><array><string>{prog}</string><string>{arg}</string></array>
  <key>StartCalendarInterval</key><dict><key>Hour</key><integer>{hour}</integer><key>Minute</key><integer>{minute}</integer></dict>
  <key>StandardOutPath</key><string>{logp}</string>
  <key>StandardErrorPath</key><string>{logp}</string>
  <key>WorkingDirectory</key><string>{HERE}</string>
</dict></plist>'''
        open(plist,"w").write(xml)
        subprocess.run(["launchctl","unload",plist],capture_output=True)
        r=subprocess.run(["launchctl","load",plist],capture_output=True,text=True)
        print(f"✅ 등록 완료(macOS launchd): 매일 {hour:02d}:{minute:02d}\n   {plist}" if r.returncode==0
              else f"❌ launchctl load 실패: {r.stderr}\n   plist 생성됨: {plist}")
    elif sysname=="Linux":
        cmd=f'{prog} "{arg}" >> "{logp}" 2>&1'
        cur=subprocess.run(["crontab","-l"],capture_output=True,text=True).stdout
        keep="\n".join(l for l in cur.splitlines() if SCHED_LABEL not in l)
        new=(keep.strip()+"\n" if keep.strip() else "")+f"{minute} {hour} * * * {cmd}  # {SCHED_LABEL}\n"
        p=subprocess.run(["crontab","-"],input=new,text=True)
        print(f"✅ 등록 완료(cron): 매일 {hour:02d}:{minute:02d}" if p.returncode==0 else "❌ crontab 등록 실패")
    elif sysname=="Windows":
        # 작업 스케줄러(schtasks) 등록. 래퍼가 .bat 이면 cmd /c 로, .py 면 python 으로 실행.
        if prog=="cmd.exe": tr=f'cmd.exe /c ""{arg}" >> "{logp}" 2>&1"'
        else:               tr=f'cmd.exe /c ""{prog}" "{arg}" >> "{logp}" 2>&1"'
        r=subprocess.run(["schtasks","/Create","/F","/SC","DAILY","/TN",SCHED_LABEL,
                          "/ST",f"{hour:02d}:{minute:02d}","/TR",tr],
                         capture_output=True,text=True)
        print(f"✅ 등록 완료(Windows 작업 스케줄러): 매일 {hour:02d}:{minute:02d}\n   작업 이름: {SCHED_LABEL}"
              if r.returncode==0 else
              f"❌ schtasks 등록 실패: {(r.stderr or r.stdout).strip()[:300]}\n"
              f"   → 관리자 권한 명령 프롬프트에서 다시 실행하거나, 작업 스케줄러 GUI 로 아래를 등록하세요:\n   {tr}")
    else:
        print(f"이 OS({sysname})는 자동 등록 미지원 — 수동으로 등록하세요.")

def uninstall_schedule():
    sysname=platform.system()
    if sysname=="Darwin":
        plist=os.path.expanduser(f"~/Library/LaunchAgents/{SCHED_LABEL}.plist")
        subprocess.run(["launchctl","unload",plist],capture_output=True)
        if os.path.exists(plist): os.remove(plist)
        print("✅ 자동 실행 해제 완료(macOS launchd)")
    elif sysname=="Linux":
        cur=subprocess.run(["crontab","-l"],capture_output=True,text=True).stdout
        keep="\n".join(l for l in cur.splitlines() if SCHED_LABEL not in l)
        subprocess.run(["crontab","-"],input=keep+"\n",text=True)
        print("✅ 자동 실행 해제 완료(cron)")
    elif sysname=="Windows":
        r=subprocess.run(["schtasks","/Delete","/F","/TN",SCHED_LABEL],capture_output=True,text=True)
        print("✅ 자동 실행 해제 완료(Windows 작업 스케줄러)" if r.returncode==0
              else f"❌ schtasks 삭제 실패(이미 없을 수 있음): {(r.stderr or r.stdout).strip()[:200]}")
    else:
        print(f"이 OS({sysname})는 자동 해제 미지원.")

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--config",default=os.path.join(HERE,"config.ini"))
    ap.add_argument("--dry-run",action="store_true",help="엑셀만 생성(발송 X)")
    ap.add_argument("--test-db",action="store_true",help="DB 접속만 확인")
    ap.add_argument("--test-mail",action="store_true",help="DB 없이 SMTP 전송만 점검")
    ap.add_argument("--setup",action="store_true",help="비밀번호 1회 입력·저장(이후 무인 실행)")
    ap.add_argument("--doctor",action="store_true",help="라이브러리/템플릿/config/DB/SMTP 일괄 점검")
    ap.add_argument("--install-schedule",action="store_true",help="매일 08:00 자동 실행 등록")
    ap.add_argument("--uninstall-schedule",action="store_true",help="자동 실행 해제")
    ap.add_argument("--date",help="특정 날짜로 생성·발송 (YYYY-MM-DD)")
    ap.add_argument("--demo",action="store_true",help="DB 없이 샘플 데이터로 리포트 생성+메일 발송(전 과정 시연)")
    ap.add_argument("--version",action="store_true",help="버전 표시")
    a=ap.parse_args()
    if a.version: print(f"CS-MES balance_outgoing_mailer v{VERSION}"); return
    if a.install_schedule:   install_schedule();   return
    if a.uninstall_schedule: uninstall_schedule(); return
    cfg=load_config(a.config)
    if a.setup:
        ensure_password(cfg,a.config,"smtp","SMTP(메일) password")
        if cfg.get("db","dsn",fallback="").strip():
            ensure_password(cfg,a.config,"db","Oracle DB password")
            if cfg.get("db","wallet_dir",fallback="").strip() and not cfg.get("db","wallet_password",fallback="").strip() and sys.stdin.isatty():
                wp=getpass.getpass("Enter wallet password (Autonomous 월렛 비번 / 없으면 Enter): ").strip()
                cfg["db"]["wallet_password"]=wp; save_password(a.config,"db",wp,key="wallet_password")
        print("설정 저장 완료. 이제 'csmes' 한 줄로 실행됩니다.")
        return
    if a.doctor: sys.exit(run_doctor(cfg,a.config))
    if a.test_mail:
        ensure_password(cfg,a.config,"smtp","SMTP(메일) password")
        try: send_test_mail(cfg)
        except Exception as e: LOG.error(f"테스트 메일 실패: {e}"); sys.exit(4)
        return
    if a.demo:
        # DB 없이 전 과정 시연: 샘플 데이터 → 실제 양식 엑셀 → 메일 발송
        before=cfg.getint("report","window_before",fallback=3); after=cfg.getint("report","window_after",fallback=7)
        today=site_today(cfg); today_str=today.strftime("%Y-%m-%d")
        buckets=build_buckets(today,before,after)
        data=make_demo_data(buckets)
        body="[DEMO 데이터] DB 미연동 상태에서 전체 흐름(리포트 생성→메일 발송)을 시연합니다.\n\n"+build_body_summary(data)
        wb=build_workbook(data,buckets,today,today_str)
        out=os.path.join(report_dir(cfg),f"BALANCE_OUTGOING_DEMO_{today.strftime('%Y%m%d')}.xlsx"); wb.save(out)
        LOG.info(f"DEMO Excel 생성: {out}"); print(body)
        if a.dry_run: print("\n--dry-run: 발송 생략 (엑셀만 생성)"); return
        ensure_password(cfg,a.config,"smtp","SMTP(메일) password")
        try: retry(lambda: send_mail(cfg,out,body,today_str+" (DEMO)"), tries=3, delay=10, label="DEMO 메일 발송")
        except Exception as e: LOG.error(f"DEMO 메일 발송 실패: {e}"); sys.exit(3)
        print("OK: DEMO 리포트를 메일로 보냈습니다. 받은편지함을 확인하세요.")
        return
    # 날짜 결정 (--date 우선)
    if a.date:
        try: today=datetime.datetime.strptime(a.date,"%Y-%m-%d").date()
        except ValueError: sys.exit("[옵션오류] --date 형식은 YYYY-MM-DD")
    else:
        today=site_today(cfg)
    today_str=today.strftime("%Y-%m-%d")
    before=cfg.getint("report","window_before",fallback=3); after=cfg.getint("report","window_after",fallback=7)
    plants=[x.strip() for x in cfg.get("report","plants").split(",") if x.strip()]
    buckets=build_buckets(today,before,after)
    d_from=min(b[1] for b in buckets); d_to=max(b[1] for b in buckets)
    LOG.info(f"=== 시작 v{VERSION} {today_str} 창 {d_from}~{d_to} plants={plants} ===")
    ensure_password(cfg,a.config,"db","Oracle DB password")
    if not a.dry_run: ensure_password(cfg,a.config,"smtp","SMTP(메일) password")
    notify = not (a.dry_run or a.test_db)   # 실패 알림 발송 여부
    # --- DB 접속(재시도) ---
    try:
        conn=retry(lambda: db_connect(cfg), tries=3, delay=5, label="DB 접속")
    except Exception as e:
        LOG.error(f"DB 접속 실패: {e}")
        if notify: send_failure_mail(cfg,f"DB 접속 실패: {e}",today_str)
        sys.exit(2)
    if a.test_db:
        try:
            cur=conn.cursor(); cur.execute("SELECT 1 FROM DUAL"); print("DB OK:",cur.fetchone())
        finally: conn.close()
        return
    # --- 조회·집계 ---
    strict=cfg.getboolean("report","strict_outgoing",fallback=True)   # 정식 GMES 로직(엄격) / OCI 동기화 전제
    LOG.info(f"미출고 판정: {'정식(strict)' if strict else '임시(loose)'}")
    try:
        data={}
        sm=fetch_scan_bembep(conn,plants,d_from,d_to)
        for name,fams in SHEETS:
            pv=pivot(fetch_sheet(conn,fams,plants,d_from,d_to,strict=strict),buckets,sm); data[name]=pv
            LOG.info(f"{name}: {len(pv)}행 잔량 {sum(r['total'] for r in pv):,}")
    except Exception as e:
        LOG.error(f"조회/집계 실패: {e}")
        if notify: send_failure_mail(cfg,f"조회/집계 실패: {e}",today_str)
        sys.exit(5)
    finally:
        try: conn.close()
        except Exception: pass
    body=build_body_summary(data)
    # --- 엑셀 생성 ---
    try:
        wb=build_workbook(data,buckets,today,today_str)
        out=os.path.join(report_dir(cfg),f"BALANCE_OUTGOING_{today.strftime('%Y%m%d')}.xlsx"); wb.save(out)
    except Exception as e:
        LOG.error(f"엑셀 생성 실패: {e}")
        if notify: send_failure_mail(cfg,f"엑셀 생성 실패: {e}",today_str)
        sys.exit(6)
    LOG.info(f"Excel 생성: {out}"); print(body)
    if a.dry_run: LOG.info("--dry-run: 발송 생략"); return
    # --- 메일 발송(재시도) ---
    try:
        retry(lambda: send_mail(cfg,out,body,today_str), tries=3, delay=10, label="메일 발송")
    except Exception as e:
        LOG.error(f"메일 발송 실패: {e}"); send_failure_mail(cfg,f"메일 발송 실패: {e}",today_str); sys.exit(3)
    LOG.info("=== 완료 ===")

if __name__=="__main__": main()
