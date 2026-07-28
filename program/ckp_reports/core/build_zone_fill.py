#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""No.2 IP Production 존(zone) 양식 — 종합 템플릿을 서식째 복사하고 데이터를 채운다 (Option A).
사용: python build_zone_fill.py OUT.xlsx SRC_종합.xlsx no2.csv 2026-07-13
  · 종합의 '3-1. Balance IP Production' 시트 서식(색·존범례·병합·JJ/RJ)을 100% 유지.
  · 날짜열은 기준일 기준 D-Day~D-7(S~Z)로 갱신, D+ 열(L~R)은 숨김.
  · 채우는 칸: Item Class(B)·LINE(C)·MODEL(D)·Style(E)·color(F)·날짜(S~Z)·PROD Total(AE).
  · 빈칸(소스 미확정): PLANT(A)·SPRAY(G)·PAD(H)·MCS(I)·YIELD(J)·Gen(K)·TOTAL KG(AF)~.
"""
import sys, csv, datetime
from copy import copy
from collections import OrderedDict, defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.formatting import ConditionalFormattingList

SHEET = "3-1. Balance IP Production"
FIRST = 10                       # 데이터 시작행
C_ITEM, C_LINE, C_MODEL, C_STYLE, C_COLOR, C_PRODTOT = 2, 3, 4, 5, 6, 31   # B C D E F AE
C_MCS, C_GEN = 9, 11             # I=MCS, K=Gen
DCOL_FROM, DCOL_TO = 19, 26      # S~Z = D-Day~D-7 (8열)
DPLUS_FROM, DPLUS_TO = 12, 18    # L~R = D+ (숨김)
STYLE_ROW = 11                   # 서식 참조용 데이터행

def buckets_dday(day, after=7):
    out = [("D-Day", day.strftime("%Y%m%d"))]
    d, n = day, 0
    while n < after:
        d += datetime.timedelta(days=1)
        if d.weekday() != 6:
            n += 1; out.append((f"D-{n}", d.strftime("%Y%m%d")))
    return out  # 8개

def read_rows(path):
    rows = []
    for d in csv.DictReader(open(path, encoding="utf-8")):
        d = {(k or "").strip().upper(): v for k, v in d.items()}
        try: q = int(float(d.get("QTY") or 0))
        except (TypeError, ValueError): q = 0
        if not q: continue
        rows.append((d.get("ITEM_CLASS", "").strip(),
                     (d.get("FA_WC") or d.get("FA_WC_CD") or "").strip(),
                     d.get("STYLE_CD", "").strip(),
                     (d.get("STYLE_NAME") or d.get("MODEL_NAME") or "").strip(),
                     d.get("MCS_COLOR", "").strip(),
                     (d.get("GEN") or "").strip(), (d.get("MCS") or "").strip(),
                     str(d.get("FA_DATE") or "").strip(),
                     (d.get("FAPLANT") or "").strip().upper(), q))
    return rows

def build(out, src_wb, csvpath, date):
    day = datetime.datetime.strptime(date, "%Y-%m-%d").date()
    bk = buckets_dday(day, 7)
    wb = openpyxl.load_workbook(src_wb)
    for s in list(wb.sheetnames):
        if s != SHEET: wb.remove(wb[s])
    ws = wb[SHEET]
    maxr, maxc = ws.max_row, ws.max_column

    # 1) 날짜 헤더(행7, S~Z) + 참조일(E4) + 월 라벨(L6) 갱신
    ws.cell(4, 5, day.strftime("%Y-%m-%d"))
    for i, (lab, ymd) in enumerate(bk):
        ws.cell(7, DCOL_FROM + i, f"{int(ymd[4:6])}-{int(ymd[6:8])}")
    # 월 라벨(L6, 병합 L6:W6) — 하드코딩 'APRIL' → 기준일의 실제 월 영어명으로
    MONTHS = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
              "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]
    ws.cell(6, 12, MONTHS[day.month - 1])

    # 2) D+ 열(L~R) 숨김
    for c in range(DPLUS_FROM, DPLUS_TO + 1):
        ws.column_dimensions[get_column_letter(c)].hidden = True

    # 3) 기존 데이터행 병합해제 + 서식 샘플 확보 + 값 클리어
    for mr in [str(m) for m in list(ws.merged_cells.ranges) if m.min_row >= FIRST]:
        ws.unmerge_cells(mr)
    styles = {c: copy(ws.cell(STYLE_ROW, c)._style) for c in range(1, maxc + 1)}
    for r in range(FIRST, maxr + 1):
        for c in range(1, maxc + 1):
            ws.cell(r, c).value = None

    # 4) 데이터 채우기 — 스타일 1개 = JJ행(위) + RJ행(아래). 설명열(A~K)은 2행 세로병합 1회 표시.
    DESC_COLS = range(1, C_GEN + 1)      # A(1)~K(11): PLANT·Item·LINE·MODEL·Style·color·SPRAY·PAD·MCS·YIELD·Gen
    grouped = OrderedDict()               # key(설명) → {"JJ": {fa:qty}, "RJ": {fa:qty}}
    for ic, wc, style, name, color, gen, mcs, fa, faplant, q in read_rows(csvpath):
        g = grouped.setdefault((ic, wc, style, name, color, gen, mcs),
                               {"JJ": defaultdict(int), "RJ": defaultdict(int)})
        g["RJ" if faplant == "RJ" else "JJ"][fa] += q
    ordered = sorted(grouped.items(), key=lambda kv: (kv[0][1], kv[0][2]))
    bdates = [ymd for _, ymd in bk]
    rr, gtot = FIRST, 0
    for (k, planes) in ordered:
        ic, wc, style, name, color, gen, mcs = k
        rJJ, rRJ = rr, rr + 1
        for c in range(1, maxc + 1):     # 두 행 모두 샘플행 서식 복제(병합 전에)
            ws.cell(rJJ, c)._style = copy(styles[c])
            ws.cell(rRJ, c)._style = copy(styles[c])
        # 설명열: 값은 위(JJ)행에, 셀은 JJ·RJ 2행 세로병합
        ws.cell(rJJ, C_ITEM, ic); ws.cell(rJJ, C_LINE, wc); ws.cell(rJJ, C_MODEL, name)
        ws.cell(rJJ, C_STYLE, style); ws.cell(rJJ, C_COLOR, color)
        if mcs: ws.cell(rJJ, C_MCS, mcs)
        if gen: ws.cell(rJJ, C_GEN, gen)
        for c in DESC_COLS:
            ws.merge_cells(start_row=rJJ, start_column=c, end_row=rRJ, end_column=c)
        # 날짜·PROD Total: 위=JJ / 아래=RJ 각각 (RJ 데이터 없어도 빈 행 유지)
        for row, plane in ((rJJ, "JJ"), (rRJ, "RJ")):
            dm = planes[plane]
            tot = sum(dm.get(d, 0) for d in bdates); gtot += tot
            for i, d in enumerate(bdates):
                v = dm.get(d, 0)
                if v: ws.cell(row, DCOL_FROM + i, v)
            ws.cell(row, C_PRODTOT, tot)
        rr += 2

    # 안 쓰는 남은 템플릿 행(데이터 아래 빈 행) 삭제 → 실제 채운 행만 남김.
    #   원본 잔재 조건부서식(값기반 하이라이트, 존 색 아님) 먼저 제거해야 행 삭제 시 범위 어긋남이 없다.
    ws.conditional_formatting = ConditionalFormattingList()
    if rr <= maxr:
        ws.delete_rows(rr, maxr - rr + 1)
    ws.auto_filter.ref = None            # 행 삭제로 어긋난 자동필터 범위 제거

    ws.print_area = None
    # 종합 템플릿에 낀 쓰레기 제거 → Excel '복구/손상' 경고 방지:
    #   ① 죽은 외부링크(externalLink*.xml, 옛 Nike 네트워크/SharePoint 참조) — orphan.
    #   ② 정의된 이름 255개(대부분 IF(#REF!) · [N]Tables!#REF! 외부링크 참조) —
    #      ①을 지우면 이 이름들이 dangling 되어 오히려 손상 경고가 뜨므로 함께 제거.
    wb._external_links = []
    wb.defined_names.clear()
    wb.save(out)
    print(f"SAVED: {out}  (zone-fill) styles={len(ordered)} rows={len(ordered)*2}(JJ/RJ) Total={gtot:,}")

if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
