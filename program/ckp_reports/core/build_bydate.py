#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CKP Manual Report — Balance ... (by date) 생성기  (3-1 계열: No.7·8·9·10)
=========================================================================
원본 3-1 시트 양식을 그대로 재현한다.
  r1 타이틀 / r2 고정8 + 날짜(MM-DD) / r3 D-offset(D+N…D-Day…D-7) [+KET] / r4~ 데이터
★ 날짜 컬럼은 **원본과 동일한 D-offset 고정 구조**로 항상 렌더된다(데이터가 없어도 열이 사라지지 않음).
  날짜는 기준일(--date)에서 작업일(일요일 제외) 기준으로 재계산.

고정8: Item Class | FA W/C | Style CD | Style Name | MCS Color | SPRAY | PAD | PROD Total
(SPRAY/PAD 는 소스 미확정으로 공란)

입력 tidy CSV: ITEM_CLASS,FA_WC,STYLE_CD,STYLE_NAME,MCS_COLOR,FA_DATE,QTY
사용:
  python balance_bydate.py OUT.xlsx "SHEET=csv" --date 2026-07-07 --before 6 --after 7 --title "BALANCE CMP" [--ket]
"""
import sys, csv, io, datetime
from collections import OrderedDict, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FIXED = ["Item Class", "FA W/C", "Style CD", "Style Name", "MCS Color", "SPRAY", "PAD", "PROD Total"]
NAVY="808080"; ALT="EAF1F8"; GT="D9E1F2"   # 헤더 회색(종합 No.7~10 동일: theme0 tint-0.5)
thin=Side(style="thin", color="BBBBBB"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(sz=9,b=False,c="222222"): return Font(name="Malgun Gothic", size=sz, bold=b, color=c)
HDR=F(9,True,"FFFFFF")
# 날짜열 D-offset 존 색 (이미지 양식): D+ 블랙 / D-Day~D-2 레드 / D-3,4 옐로우 / D-5,6 그린 / D-7 무색
ZRED="FF0000"; ZYEL="FFFF00"; ZGRN="92D050"; ZBLK="000000"
def _zone(label):
    if label.startswith("D+"):         return (ZBLK, "FFFFFF")
    if label in ("D-Day","D-1","D-2"): return (ZRED, "FFFFFF")
    if label in ("D-3","D-4"):         return (ZYEL, "000000")
    if label in ("D-5","D-6"):         return (ZGRN, "000000")
    return (None, None)                # D-7 등 = 무색

def _num(v):
    try: return int(float(v))
    except (TypeError, ValueError): return None

def build_buckets(day, before, after):
    """원본과 동일: 작업일(일요일 제외) 기준 D+before … D-Day … D-after. 반환 [(label, YYYYMMDD)]."""
    def walk(n, step):
        out=[]; d=day
        while len(out) < n:
            d = d + datetime.timedelta(days=step)
            if d.weekday() != 6: out.append(d)
        return out
    seq = list(reversed(walk(before, -1))) + [day] + walk(after, +1)
    labels = [f"D+{k}" for k in range(before,0,-1)] + ["D-Day"] + [f"D-{k}" for k in range(1,after+1)]
    return [(labels[i], seq[i].strftime("%Y%m%d")) for i in range(len(seq))]

def read_tidy(path):
    with open(path, encoding="utf-8") as f: txt=f.read()
    rows=[]
    for d in csv.DictReader(io.StringIO(txt)):
        d={(k or "").strip().upper(): v for k,v in d.items()}
        q=_num(d.get("QTY"))
        if not q: continue
        rows.append({"ic":(d.get("ITEM_CLASS") or "").strip(),
                     "wc":(d.get("FA_WC") or d.get("FA_WC_CD") or "").strip(),
                     "style":(d.get("STYLE_CD") or "").strip(),
                     "name":(d.get("STYLE_NAME") or d.get("MODEL_NAME") or "").strip(),
                     "color":(d.get("MCS_COLOR") or "").strip(),
                     "fa":str(d.get("FA_DATE") or "").strip(), "qty":q})
    return rows

def build_sheet(ws, rows, buckets, title, ket=False):
    nb=len(buckets); NC=8+nb+(1 if ket else 0)
    ws.cell(1,1,title).font=F(11,True,NAVY)
    # r2: 고정8 + 날짜 / r3: D-offset
    for c,h in enumerate(FIXED,1):
        x=ws.cell(2,c,h); x.font=HDR; x.fill=PatternFill("solid",fgColor=NAVY)
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=BORDER
        ws.merge_cells(start_row=2,start_column=c,end_row=3,end_column=c)
        ws.cell(3,c).fill=PatternFill("solid",fgColor=NAVY); ws.cell(3,c).border=BORDER
    for i,(lab,d) in enumerate(buckets):
        c=9+i
        x=ws.cell(2,c,f"{d[4:6]}-{d[6:8]}"); x.font=HDR; x.fill=PatternFill("solid",fgColor=NAVY)
        x.alignment=Alignment(horizontal="center"); x.border=BORDER
        zf,zc=_zone(lab)                                   # r3 D-offset: D+·D-7 회색, 나머지 존색
        hf,hc = (zf,zc) if (zf and not lab.startswith("D+")) else (NAVY,"FFFFFF")
        y=ws.cell(3,c,lab); y.font=F(8,True,hc); y.fill=PatternFill("solid",fgColor=hf)
        y.alignment=Alignment(horizontal="center"); y.border=BORDER
    if ket:
        c=9+nb
        for r in (2,3):
            x=ws.cell(r,c,"KET" if r==2 else None); x.font=HDR; x.fill=PatternFill("solid",fgColor=NAVY); x.border=BORDER
        ws.merge_cells(start_row=2,start_column=c,end_row=3,end_column=c)

    bdate=[d for _,d in buckets]
    key=lambda r:(r["ic"],r["wc"],r["style"],r["name"],r["color"])
    grouped=OrderedDict()
    for r in rows: grouped.setdefault(key(r), defaultdict(int))[r["fa"]] += r["qty"]
    ordered=sorted(grouped.items(), key=lambda kv:(kv[0][1], kv[0][0], kv[0][2]))

    rr=4; alt=0
    for (k,dm) in ordered:
        ic,wc,style,name,color = k
        tot=sum(dm.get(d,0) for d in bdate); alt^=1
        vals=[ic,wc,style,name,color,"","",tot] + [ (dm.get(d,0) or None) for d in bdate ]
        if ket: vals.append(None)
        for c,v in enumerate(vals,1):
            x=ws.cell(rr,c,v); x.font=F(9); x.border=BORDER
            if 9 <= c < 9+nb:                       # 날짜 열: 값 있으면 존 색, 없으면 무색(흰색)
                zf,zc=_zone(buckets[c-9][0])
                if v is not None and zf:
                    x.fill=PatternFill("solid",fgColor=zf); x.font=F(9,True,zc)
            elif alt:
                x.fill=PatternFill("solid",fgColor=ALT)
        rr+=1
    # 데이터가 없어도 헤더/열은 그대로 유지된다.
    widths=[10,8,14,24,18,7,6,10]+[7]*nb+([8] if ket else [])
    for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="I4"; ws.sheet_view.showGridLines=False
    ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=9
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    return sum(sum(dm.get(d,0) for d in bdate) for _,dm in ordered), len(ordered)

def main():
    a=list(sys.argv[1:])
    def pop(flag, default=None):
        if flag in a:
            i=a.index(flag); v=a[i+1]; del a[i:i+2]; return v
        return default
    date  = pop("--date", datetime.date.today().isoformat())
    before= int(pop("--before", "6")); after = int(pop("--after", "7"))
    title = pop("--title", "BALANCE")
    ket   = "--ket" in a
    if ket: a.remove("--ket")
    out=a[0]; specs=a[1:]
    day=datetime.datetime.strptime(date,"%Y-%m-%d").date()
    buckets=build_buckets(day, before, after)
    wb=openpyxl.Workbook(); first=True; summ=[]
    for spec in specs:
        name,path=spec.split("=",1); rows=read_tidy(path)
        ws=wb.active if first else wb.create_sheet(); first=False; ws.title=name[:31]
        tot,nrow=build_sheet(ws, rows, buckets, f"{title} {date}", ket)
        summ.append(f"· {name}: {nrow}행 / Total {tot:,} / 날짜열 {len(buckets)}개")
    wb.save(out); print("SAVED:",out); print("\n".join(summ))

if __name__=="__main__": main()
